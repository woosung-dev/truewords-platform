"""Phase 4 3-way Codex 정성 평가 markdown 생성.

3개 컬렉션 (Recursive baseline / Hierarchical / Contextual) 비교 사례를 추출.
L별 stratified sampling, 동일 질문에 대한 3개 답변을 한 번에 비교.

사용:
    PYTHONPATH=. uv run python scripts/build_phase4_3way_codex.py \\
        --xlsx-recursive ../tmp_match/phase4_eval/eval_recursive_*.xlsx \\
        --xlsx-hierarchical ../tmp_match/phase4_eval/eval_hierarchical_*.xlsx \\
        --xlsx-contextual ../tmp_match/phase4_eval/eval_contextual_*.xlsx \\
        --output ../tmp_match/phase4_eval/codex_compare_3way.md
"""
from __future__ import annotations

import argparse
import random
from pathlib import Path

from openpyxl import load_workbook


def detect_level(level_raw: str) -> str:
    for L in ("L1", "L2", "L3", "L4", "L5"):
        if L in level_raw:
            return L
    return "기타"


def find_answer_col(headers: list[str]) -> str | None:
    for h in headers:
        if "우리 답변" in h:
            return h
    return None


def load_rows(path: Path) -> tuple[list[dict], str]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], ""
    headers = [str(h) if h is not None else "" for h in rows[0]]
    ans_col = find_answer_col(headers) or ""
    out = []
    for r in rows[1:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        out.append({headers[i]: r[i] if i < len(r) else None for i in range(len(headers))})
    return out, ans_col


def render_case(idx: int, q_row: dict, ans_r: str, ctx_r: list[str],
                ans_h: str, ctx_h: list[str], ans_c: str, ctx_c: list[str]) -> list[str]:
    L = detect_level(str(q_row.get("난이도(Level)", "")))
    cat = str(q_row.get("카테고리", "") or "")
    q = (q_row.get("테스트용 질문") or "").strip()
    gt = str(q_row.get("봇 모범 답변", "") or "")
    kw = str(q_row.get("참고 키워드", "") or "")

    out = [
        f"## 사례 {idx} — {L} ({cat})",
        "",
        f"**질문**: {q}",
        "",
        f"**모범답변**: {gt}",
        "",
        f"**참고 키워드**: {kw}",
        "",
        "### 🔵 Recursive (baseline)",
        ans_r,
        "",
    ]
    if ctx_r:
        out.append("**참고**:")
        for i, c in enumerate(ctx_r, 1):
            out.append(f"{i}. {c[:280]}")
        out.append("")

    out.extend([
        "### 🟢 Hierarchical (Parent-Child)",
        ans_h,
        "",
    ])
    if ctx_h:
        out.append("**참고**:")
        for i, c in enumerate(ctx_h, 1):
            out.append(f"{i}. {c[:280]}")
        out.append("")

    out.extend([
        "### 🟣 Contextual Retrieval (Anthropic)",
        ans_c,
        "",
    ])
    if ctx_c:
        out.append("**참고**:")
        for i, c in enumerate(ctx_c, 1):
            out.append(f"{i}. {c[:280]}")
        out.append("")

    out.append("---")
    out.append("")
    return out


def main() -> int:
    p = argparse.ArgumentParser()
    p.add_argument("--xlsx-recursive", required=True, type=Path)
    p.add_argument("--xlsx-hierarchical", required=True, type=Path)
    p.add_argument("--xlsx-contextual", required=True, type=Path)
    p.add_argument("--output", required=True, type=Path)
    p.add_argument("--per-level", type=int, default=2)
    p.add_argument("--seed", type=int, default=20260501)
    args = p.parse_args()

    rows_r, ans_r_col = load_rows(args.xlsx_recursive)
    rows_h, ans_h_col = load_rows(args.xlsx_hierarchical)
    rows_c, ans_c_col = load_rows(args.xlsx_contextual)

    by_q_h = {(r.get("테스트용 질문") or "").strip(): r for r in rows_h}
    by_q_c = {(r.get("테스트용 질문") or "").strip(): r for r in rows_c}

    # L별 stratify (Recursive 기준)
    rng = random.Random(args.seed)
    by_level: dict[str, list[dict]] = {}
    for r in rows_r:
        L = detect_level(str(r.get("난이도(Level)", "")))
        if L in ("L1", "L2", "L3", "L4", "L5"):
            by_level.setdefault(L, []).append(r)

    selected: list[dict] = []
    for L in ("L1", "L2", "L3", "L4", "L5"):
        bucket = by_level.get(L, [])
        n = min(args.per_level, len(bucket))
        if n:
            selected.extend(rng.sample(bucket, n))

    lines = [
        "# Phase 4 Codex 3-way 정성 검토 — Recursive vs Hierarchical vs Contextual",
        "",
        "## 평가 요청",
        "",
        f"5권 신학/원리 PoC scope, 통일원리 100선 평가셋 — {len(selected)}건 sampling.",
        "",
        "**자동 메트릭 (LLM-Judge 4 metric × 1~5점, 총점 4~20)**:",
        "",
        "| 컬렉션 | 총점 | L1 | L2 | L3 | L4 | L5 | 키워드 F1 |",
        "|---|---:|---:|---:|---:|---:|---:|---:|",
        "| 🔵 Recursive (baseline) | **16.93** | 18.65 | 18.40 | 15.35 | 15.95 | 16.30 | 0.611 |",
        "| 🟢 Hierarchical | 16.38 | 18.05 | 17.90 | 14.05 | 15.55 | 16.35 | 0.606 |",
        "| 🟣 Contextual | 16.29 | 17.70 | 17.05 | 15.00 | 15.50 | 16.20 | 0.615 |",
        "| **Δ vs Recursive** | | | | | | | |",
        "| Hierarchical | -0.55 | -0.60 | -0.50 | -1.30 | -0.40 | +0.05 | -0.005 |",
        "| Contextual | -0.64 | -0.95 | -1.35 | -0.35 | -0.45 | -0.10 | +0.004 |",
        "",
        "→ **자동 메트릭은 Recursive 우월** (둘 다 임계값 +0.5 미달).",
        "",
        "다음 사례들에서 정성적 평가를 부탁드립니다:",
        "1. **각 사례 승자**: Recursive / Hierarchical / Contextual / 동등",
        "2. **이유**: 모범답변/참고키워드 대비 정확도, 컨텍스트 활용도, 환각 여부",
        "",
        "마지막에 종합 의견:",
        "- 사례별 승자 집계",
        "- 메트릭별(정확도/문맥/환각) 우열 패턴",
        "- L별(L1~L5) 패턴",
        "- **운영 적용 권장 chunking 전략** + 보강할 약점",
        "- 후속 PR 우선순위 (88권 재임베딩 가치 vs v5 Recursive 유지)",
        "",
        "**참고 — Phase 4 PoC 비용/시간**:",
        "- Hierarchical 5권: 27,074 청크 (Recursive 10,558의 2.6배), 28분 적재",
        "- Contextual 5권: 10,558 청크 + Gemini Flash Lite 컨텍스트 prefix 생성, 21분 적재",
        "- 88권 확장 시: Hierarchical ~₩600 / Contextual ~₩2,000+ + prompt caching 비용",
        "",
        "---",
        "",
    ]

    for i, r in enumerate(selected, 1):
        q = (r.get("테스트용 질문") or "").strip()
        rh = by_q_h.get(q, {})
        rc = by_q_c.get(q, {})

        ans_r = str(r.get(ans_r_col, "") or "")
        ans_h = str(rh.get(ans_h_col, "") or "(매칭 실패)")
        ans_c = str(rc.get(ans_c_col, "") or "(매칭 실패)")

        ctx_r = [str(r.get(col, "") or "") for col in ("참고1", "참고2", "참고3") if r.get(col)]
        ctx_h = [str(rh.get(col, "") or "") for col in ("참고1", "참고2", "참고3") if rh.get(col)]
        ctx_c = [str(rc.get(col, "") or "") for col in ("참고1", "참고2", "참고3") if rc.get(col)]

        lines.extend(render_case(i, r, ans_r, ctx_r, ans_h, ctx_h, ans_c, ctx_c))

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text("\n".join(lines), encoding="utf-8")
    print(f"✓ Codex 3-way 비교 markdown {len(selected)}건 생성: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
