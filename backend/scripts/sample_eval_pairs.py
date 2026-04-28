"""
RAGAS 평가용 Stratified 50건 샘플러.

200건 평가 xlsx 3개에서 (파일 × 난이도) 비례로 50건을 뽑고,
참고1/2/3 셀에서 contexts(본문 텍스트)만 추출해 RAGAS 호환 포맷으로 출력한다.

입력 xlsx 컬럼 (3파일 모두 동일):
    번호, 난이도(Level), 카테고리, 테스트용 질문, 봇 모범 답변,
    참고 키워드, 우리 답변(all), 참고1, 참고2, 참고3, 세션ID

참고1/2/3 셀 포맷:
    [문서명.txt] (score=0.750, source=N)
    <본문 텍스트>

출력:
    - xlsx (사람 검수용): id, source_file, level, category, question,
      ground_truth, our_answer, context_1, context_2, context_3,
      session_id, n_contexts
    - json (eval_ragas.py 입력용): [{"id", "source_file", "level",
      "category", "question", "answer", "ground_truth",
      "contexts": list[str]}, ...]

사용:
    cd backend
    uv run --group eval python scripts/sample_eval_pairs.py \\
        --output-dir ~/Downloads
"""

from __future__ import annotations

import argparse
import json
import random
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

DEFAULT_INPUT_DIR = Path.home() / "Downloads"

# 파일별 난이도 quota (총 50건). 행 분포 비례.
ALLOCATION: dict[str, dict[str, int]] = {
    "notebooklm_qa_전체검색봇_평가_튜닝후_20260427_1649.xlsx": {
        "L1 (단순 사실 조회)": 5,
        "L2 (출처 인용)": 5,
        "L3 (주제 요약)": 5,
        "L4 (개념 연결)": 5,
        "L5 (교리 추론)": 5,
    },
    "천일국섭리_튜닝후.xlsx": {
        "긍정적": 3,
        "평균적": 5,
        "부정적": 4,
    },
    "참부모섭리_튜닝후.xlsx": {
        "간단": 5,
        "보통": 5,
        "상세": 3,
    },
}

# `[문서명] (score=0.750, source=N)\n<본문>` 형식에서 metadata + 본문 분리
CONTEXT_PATTERN = re.compile(
    r"^\[(?P<doc>[^\]]+)\]\s*\(score=(?P<score>[-\d.]+),\s*source=(?P<source>[^)]+)\)\s*\n(?P<text>.*)$",
    re.DOTALL,
)


@dataclass
class EvalRow:
    """xlsx 한 행을 RAGAS 호환 구조로 정규화."""

    id: str
    source_file: str
    number: str
    level: str
    category: str
    question: str
    ground_truth: str
    answer: str
    contexts: list[str] = field(default_factory=list)
    raw_contexts: list[str] = field(default_factory=list)
    session_id: str = ""

    def to_json(self) -> dict[str, object]:
        return {
            "id": self.id,
            "source_file": self.source_file,
            "level": self.level,
            "category": self.category,
            "question": self.question,
            "answer": self.answer,
            "ground_truth": self.ground_truth,
            "contexts": self.contexts,
        }


def parse_context_cell(cell_value: str | None) -> str | None:
    """참고N 셀에서 본문(text)만 추출. 형식이 깨졌거나 비어 있으면 None."""
    if not cell_value:
        return None
    text = str(cell_value).strip()
    if not text:
        return None
    match = CONTEXT_PATTERN.match(text)
    if match:
        body = match.group("text").strip()
        return body or None
    # 패턴이 안 맞으면 셀 전체를 그대로 사용 (보수적)
    return text


def load_rows(path: Path) -> list[EvalRow]:
    """xlsx의 행을 EvalRow 리스트로 로드."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"활성 시트 없음: {path}")
    rows: list[EvalRow] = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        if not raw or not raw[0]:
            continue
        # 컬럼 인덱스: 0=번호 1=난이도 2=카테고리 3=질문 4=모범답변 5=참고키워드
        # 6=우리답변 7=참고1 8=참고2 9=참고3 10=세션ID
        number = str(raw[0])
        level = str(raw[1]) if raw[1] else ""
        category = str(raw[2]) if raw[2] else ""
        question = str(raw[3]) if raw[3] else ""
        ground_truth = str(raw[4]) if raw[4] else ""
        answer = str(raw[6]) if len(raw) > 6 and raw[6] else ""
        raw_contexts = [str(raw[i]) if len(raw) > i and raw[i] else "" for i in (7, 8, 9)]
        contexts = [c for c in (parse_context_cell(rc) for rc in raw_contexts) if c]
        session_id = str(raw[10]) if len(raw) > 10 and raw[10] else ""
        rows.append(
            EvalRow(
                id=f"{path.stem}#{number}",
                source_file=path.name,
                number=number,
                level=level,
                category=category,
                question=question,
                ground_truth=ground_truth,
                answer=answer,
                contexts=contexts,
                raw_contexts=raw_contexts,
                session_id=session_id,
            )
        )
    wb.close()
    return rows


def stratified_sample(
    rows_by_level: dict[str, list[EvalRow]],
    quota: dict[str, int],
    rng: random.Random,
) -> list[EvalRow]:
    """난이도별 quota만큼 랜덤 샘플링. quota를 채우지 못하면 부족분만큼 부족."""
    sampled: list[EvalRow] = []
    for level, n in quota.items():
        pool = rows_by_level.get(level, [])
        if not pool:
            print(f"  [warn] level={level!r} pool empty, skipped")
            continue
        if n >= len(pool):
            sampled.extend(pool)
            print(f"  level={level!r} 전체 {len(pool)}건 사용 (요청 {n})")
        else:
            sampled.extend(rng.sample(pool, n))
            print(f"  level={level!r} {n}건 샘플링 (pool {len(pool)})")
    return sampled


def sample_all(input_dir: Path, seed: int) -> list[EvalRow]:
    rng = random.Random(seed)
    all_sampled: list[EvalRow] = []
    for filename, quota in ALLOCATION.items():
        path = input_dir / filename
        if not path.exists():
            raise FileNotFoundError(f"입력 파일 없음: {path}")
        print(f"\n[load] {path.name}")
        rows = load_rows(path)
        rows_by_level: dict[str, list[EvalRow]] = defaultdict(list)
        for r in rows:
            rows_by_level[r.level].append(r)
        print(f"  총 {len(rows)}행, 난이도 분포: { {k: len(v) for k, v in rows_by_level.items()} }")
        sampled = stratified_sample(rows_by_level, quota, rng)
        print(f"  → 샘플 {len(sampled)}건")
        all_sampled.extend(sampled)
    return all_sampled


def write_json(rows: list[EvalRow], out_path: Path) -> None:
    payload = [r.to_json() for r in rows]
    out_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_xlsx(rows: list[EvalRow], out_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]  # 새 Workbook은 항상 active sheet 보유
    ws.title = "ragas_seed"
    headers = [
        "id",
        "source_file",
        "level",
        "category",
        "question",
        "ground_truth",
        "our_answer",
        "n_contexts",
        "context_1",
        "context_2",
        "context_3",
        "session_id",
    ]
    ws.append(headers)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        ws.append(
            [
                r.id,
                r.source_file,
                r.level,
                r.category,
                r.question,
                r.ground_truth,
                r.answer,
                len(r.contexts),
                r.contexts[0] if len(r.contexts) > 0 else "",
                r.contexts[1] if len(r.contexts) > 1 else "",
                r.contexts[2] if len(r.contexts) > 2 else "",
                r.session_id,
            ]
        )

    # 컬럼 너비 + wrap
    widths = [28, 32, 22, 18, 40, 50, 50, 10, 60, 60, 60, 36]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    wb.save(out_path)


def summarize(rows: list[EvalRow]) -> None:
    by_file = defaultdict(int)
    by_level = defaultdict(int)
    by_category = defaultdict(int)
    n_with_contexts = 0
    n_contexts_total = 0
    for r in rows:
        by_file[r.source_file] += 1
        by_level[f"{r.source_file}::{r.level}"] += 1
        by_category[r.category] += 1
        if r.contexts:
            n_with_contexts += 1
        n_contexts_total += len(r.contexts)
    print(f"\n=== Summary (총 {len(rows)}건) ===")
    print(f"파일별: {dict(by_file)}")
    print(f"파일×난이도별: {dict(by_level)}")
    print(f"카테고리별 (top 10): {dict(sorted(by_category.items(), key=lambda x: -x[1])[:10])}")
    print(f"contexts 보유: {n_with_contexts}/{len(rows)} ({100*n_with_contexts/len(rows):.0f}%)")
    print(f"평균 contexts 수: {n_contexts_total/len(rows):.2f}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help=f"입력 xlsx 디렉토리 (default: {DEFAULT_INPUT_DIR})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help=f"출력 디렉토리 (default: {DEFAULT_INPUT_DIR})",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=42,
        help="랜덤 시드 (재현성). default=42",
    )
    parser.add_argument(
        "--prefix",
        type=str,
        default="ragas_eval_seed_50",
        help="출력 파일 prefix",
    )
    args = parser.parse_args()

    sampled = sample_all(args.input_dir, args.seed)
    summarize(sampled)

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    args.output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = args.output_dir / f"{args.prefix}_{ts}.xlsx"
    json_path = args.output_dir / f"{args.prefix}_{ts}.json"
    write_xlsx(sampled, xlsx_path)
    write_json(sampled, json_path)
    print(f"\n→ {xlsx_path}")
    print(f"→ {json_path}")


if __name__ == "__main__":
    main()
