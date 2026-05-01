"""기존 측정 결과물 5개 (v1 / v2 옵션B prefix / v3 옵션F PoC / v3 옵션F 본 가동 옛 / v3 재측정 'all' 봇 토글) 통합.

LLM 호출 없음 — 기존 xlsx 파싱 + 질문 기준 join + hit rate 계산.

출력 시트:
  1. A_B 5단계 비교 — 100문항 × 5개 버전 답변 + hit 색상
  2. 요약 — 방식별 hit율 표
  3. 봇 스펙 — 각 측정 시점의 봇 spec (system_prompt 길이, tiers, persona, collection_main)

사용:
    PYTHONPATH=. uv run python scripts/merge_ab_results.py \\
        --output ~/Downloads/ab_comparison_v1_v2_v3_full.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from scripts.analyze_notebooklm_categories import _is_hit, _load_rows
from src.qdrant_client import get_client


# 측정 결과 파일 6개 (v3prodSync 추가 — 동기화된 'all-paragraph' 봇 측정)
DEFAULT_FILES = {
    "v1": "/Users/woosung/Downloads/notebooklm_post_phase1_20260428_1001_light100.xlsx",
    "v2": "/Users/woosung/Downloads/notebooklm_v2_20260428_2338_light100.xlsx",
    "v3poc": "/Users/woosung/Downloads/notebooklm_chunking_paragraph_20260429_0647.xlsx",
    "v3prodOld": "/Users/woosung/Downloads/notebooklm_prod_ab_all-paragraph_20260429_1501.xlsx",
    "v3new": "/Users/woosung/Downloads/notebooklm_v3_remeasure_20260429_1716.xlsx",
    "v3prodSync": "/Users/woosung/Downloads/notebooklm_v3prodSync_20260429_1741.xlsx",
}

LABELS = {
    "v1": "v1 baseline (sentence)",
    "v2": "v2 옵션 B (prefix, 폐기)",
    "v3poc": "v3 옵션 F PoC (평화경 1권만 paragraph, 봇 설정 불공평)",
    "v3prodOld": "v3 옵션 F 본 가동 옛 측정 (88권 paragraph, 봇 설정 불공평)",
    "v3new": "v3 재측정 ★ ('all' 봇 토글, 봇 설정 공평)",
    "v3prodSync": "v3 ★ ('all-paragraph' 봇 system_prompt 동기화 후, 봇 설정 공평)",
}

# 100건 직접 4메트릭 측정 결과 (Gemini 3.1-flash-lite-preview, langchain 우회)
DIRECT_METRICS_FILES = {
    "v1": "/Users/woosung/Downloads/metrics_direct_v1_20260429_1742.xlsx",
    "v2": "/Users/woosung/Downloads/metrics_direct_v2_20260429_1742.xlsx",
    "v3new": "/Users/woosung/Downloads/metrics_direct_v3new_20260429_1742.xlsx",
    "v3prodSync": "/Users/woosung/Downloads/metrics_direct_v3prodSync_20260429_1742.xlsx",
}

# 측정 시점에 사용된 봇 + 컬렉션 (audit 정보)
MEASUREMENT_BOTS = {
    "v1": {
        "chatbot_id": "all",
        "collection_main_at_measurement": "malssum_poc",
        "system_prompt_at_measurement": "3,607자 (정상)",
        "search_tiers_at_measurement": "weighted multi-tier (정상)",
        "persona_at_measurement": "30년 경력 가정연합 목회공직자",
        "공평성": "✅ 공평 (운영 'all' 봇 그대로)",
    },
    "v2": {
        "chatbot_id": "all",
        "collection_main_at_measurement": "malssum_poc_v2 (임시 토글)",
        "system_prompt_at_measurement": "3,607자 (정상, 'all' 봇 그대로)",
        "search_tiers_at_measurement": "weighted multi-tier ('all' 봇 그대로)",
        "persona_at_measurement": "30년 경력 가정연합 목회공직자",
        "공평성": "✅ 공평 ('all' 봇 collection_main만 토글)",
    },
    "v3poc": {
        "chatbot_id": "chunking-paragraph",
        "collection_main_at_measurement": "malssum_chunking_poc_paragraph",
        "system_prompt_at_measurement": "0자 (빈 프롬프트) ❌",
        "search_tiers_at_measurement": "단순 1-tier [A,B,C] threshold 0.6 ❌",
        "persona_at_measurement": "(빈값) ❌",
        "공평성": "❌ 불공평 (PoC 신규 봇, 'all' 봇 설정 미복사)",
    },
    "v3prodOld": {
        "chatbot_id": "all-paragraph",
        "collection_main_at_measurement": "malssum_poc_v3",
        "system_prompt_at_measurement": "0자 (빈 프롬프트) ❌",
        "search_tiers_at_measurement": "단순 1-tier [A,B,C] threshold 0.6 ❌",
        "persona_at_measurement": "(빈값) ❌",
        "공평성": "❌ 불공평 (신규 봇, 'all' 봇 설정 미복사)",
    },
    "v3new": {
        "chatbot_id": "all",
        "collection_main_at_measurement": "malssum_poc_v3 (임시 토글)",
        "system_prompt_at_measurement": "3,607자 (정상, 'all' 봇 그대로)",
        "search_tiers_at_measurement": "weighted multi-tier ('all' 봇 그대로)",
        "persona_at_measurement": "30년 경력 가정연합 목회공직자",
        "공평성": "✅ 공평 ('all' 봇 collection_main만 토글) ★",
    },
    "v3prodSync": {
        "chatbot_id": "all-paragraph",
        "collection_main_at_measurement": "malssum_poc_v3",
        "system_prompt_at_measurement": "3,607자 ('all' 봇과 md5 동일 — 동기화 완료)",
        "search_tiers_at_measurement": "weighted multi-tier ('all' 봇과 동일)",
        "persona_at_measurement": "30년 경력 가정연합 목회공직자",
        "공평성": "✅ 공평 (system_prompt 동기화 후 측정) ★",
    },
}


def _answer_col(rows: list[dict]) -> str | None:
    if not rows:
        return None
    for k in rows[0].keys():
        if "우리 답변" in k or k == "답변" or k.startswith("답변"):
            return k
    return None


def _qdrant_collection_chunk_count(name: str) -> str:
    try:
        client = get_client()
        c = client.count(name, exact=True).count
        return f"{c:,}"
    except Exception:
        return "(조회 실패)"


def main() -> int:
    parser = argparse.ArgumentParser()
    for ver, path in DEFAULT_FILES.items():
        parser.add_argument(f"--{ver}", default=path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    files = {ver: getattr(args, ver) for ver in DEFAULT_FILES}
    rows_by_ver: dict[str, list[dict]] = {}
    for ver, path in files.items():
        p = Path(path)
        if not p.exists():
            print(f"⚠️  {ver} 파일 없음: {p}", file=sys.stderr)
            rows_by_ver[ver] = []
        else:
            rows_by_ver[ver] = _load_rows(p)
            print(f"  {ver}: {len(rows_by_ver[ver])}개 ({p.name})")

    base = rows_by_ver["v1"] or rows_by_ver["v3new"]

    def question_of(r: dict) -> str:
        return (r.get("테스트용 질문") or r.get("질문") or "").strip()

    def by_question(rows: list[dict]) -> dict[str, dict]:
        return {question_of(r): r for r in rows if question_of(r)}

    indexed = {ver: by_question(rs) for ver, rs in rows_by_ver.items()}

    wb = Workbook()
    ws = wb.active
    if ws is None:
        return 1
    ws.title = "A_B 6단계 비교"

    versions = ["v1", "v2", "v3poc", "v3prodOld", "v3new", "v3prodSync"]
    header = [
        "번호", "난이도(Level)", "카테고리", "테스트용 질문", "봇 모범 답변", "참고 키워드",
    ]
    for ver in versions:
        header += [f"{ver}_답변", f"{ver}_hit"]
    ws.append(header)

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    fill_pass = PatternFill("solid", fgColor="C6EFCE")
    fill_fail = PatternFill("solid", fgColor="FFCCCC")

    for r in base:
        q = question_of(r)
        if not q:
            continue
        row_out = [
            r.get("번호", ""),
            r.get("난이도(Level)", ""),
            r.get("카테고리", ""),
            q,
            r.get("봇 모범 답변", ""),
            r.get("참고 키워드", ""),
        ]
        hit_cells: list[tuple[int, bool]] = []
        for ver in versions:
            ver_row = indexed[ver].get(q)
            if ver_row:
                ans_col = _answer_col([ver_row])
                ans = ver_row.get(ans_col, "") if ans_col else ""
                hit = _is_hit(ver_row)
            else:
                ans = ""
                hit = False
            row_out.append(ans)
            row_out.append("✅" if hit else "❌")
            hit_cells.append((len(row_out), hit))

        ws.append(row_out)
        cur_row = ws.max_row
        for col_idx, hit in hit_cells:
            ws.cell(row=cur_row, column=col_idx).fill = fill_pass if hit else fill_fail
            ws.cell(row=cur_row, column=col_idx).alignment = Alignment(horizontal="center")

    # 컬럼 너비
    base_widths = {"A": 6, "B": 12, "C": 12, "D": 50, "E": 50, "F": 25}
    for col, w in base_widths.items():
        ws.column_dimensions[col].width = w
    # 답변 컬럼 (G, I, K, M, O) + hit 컬럼 (H, J, L, N, P)
    for i, _ in enumerate(versions):
        ans_col_letter = chr(ord("G") + i * 2)
        hit_col_letter = chr(ord("G") + i * 2 + 1)
        ws.column_dimensions[ans_col_letter].width = 60
        ws.column_dimensions[hit_col_letter].width = 8

    # === 요약 시트 ===
    summary = wb.create_sheet("요약")
    summary.append(["방식", "라벨", "전체 hit", "L1", "L2", "L3", "L4", "L5", "총 문항"])
    for ver in versions:
        rs = rows_by_ver[ver]
        if not rs:
            continue
        from collections import defaultdict
        by_lvl = defaultdict(lambda: [0, 0])
        overall = [0, 0]
        for r in rs:
            lvl_raw = (r.get("난이도(Level)") or r.get("난이도") or "").strip()
            L_match = "기타"
            for L in ["L1", "L2", "L3", "L4", "L5"]:
                if L in lvl_raw:
                    L_match = L
                    break
            hit = 1 if _is_hit(r) else 0
            by_lvl[L_match][0] += hit
            by_lvl[L_match][1] += 1
            overall[0] += hit
            overall[1] += 1
        cells = []
        for L in ["L1", "L2", "L3", "L4", "L5"]:
            h, t = by_lvl[L]
            cells.append(f"{h/t:.3f}" if t else "-")
        summary.append([
            ver,
            LABELS[ver],
            f"{overall[0]/overall[1]:.3f}" if overall[1] else "-",
            *cells,
            overall[1],
        ])
    for cell in summary[1]:
        cell.font = bold
        cell.fill = header_fill
    summary.column_dimensions["A"].width = 12
    summary.column_dimensions["B"].width = 50
    for c in "CDEFGHI":
        summary.column_dimensions[c].width = 10

    # === 봇 스펙 시트 ===
    spec = wb.create_sheet("봇 스펙 (공평성 검증)")
    spec.append([
        "방식", "라벨", "측정 시점 chatbot_id", "측정 시점 collection_main",
        "system_prompt", "search_tiers", "persona", "공평성",
        "컬렉션 청크 수 (현재)",
    ])
    for ver in versions:
        info = MEASUREMENT_BOTS[ver]
        # 컬렉션 청크 수
        coll_name = info["collection_main_at_measurement"].split(" ")[0]
        chunk_count = _qdrant_collection_chunk_count(coll_name)
        spec.append([
            ver,
            LABELS[ver],
            info["chatbot_id"],
            info["collection_main_at_measurement"],
            info["system_prompt_at_measurement"],
            info["search_tiers_at_measurement"],
            info["persona_at_measurement"],
            info["공평성"],
            chunk_count,
        ])

    for cell in spec[1]:
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    spec_widths = {"A": 12, "B": 50, "C": 22, "D": 35, "E": 30, "F": 35, "G": 30, "H": 50, "I": 18}
    for col, w in spec_widths.items():
        spec.column_dimensions[col].width = w

    # 공평성 셀 색상
    fairness_pass = PatternFill("solid", fgColor="C6EFCE")
    fairness_fail = PatternFill("solid", fgColor="FFCCCC")
    for row_idx in range(2, spec.max_row + 1):
        cell = spec.cell(row=row_idx, column=8)
        text = str(cell.value or "")
        cell.fill = fairness_pass if "✅" in text else fairness_fail
        for col_idx in range(1, 10):
            spec.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    # === 4메트릭 직접 측정 시트 (langchain 우회, gemini-3.1-flash-lite-preview, n=100) ===
    metrics_summary = wb.create_sheet("4메트릭 직접 측정 (n=100)")
    metrics_summary.append([
        "방식", "라벨", "n",
        "faithfulness", "context_precision", "context_recall", "response_relevancy",
        "평균",
    ])
    metric_cols = ["faithfulness", "context_precision", "context_recall", "response_relevancy"]
    metrics_data: dict[str, dict[str, float]] = {}
    for ver, mpath in DIRECT_METRICS_FILES.items():
        mp = Path(mpath)
        if not mp.exists():
            continue
        mwb = load_workbook(mp, data_only=True)
        mws = mwb["metrics"] if "metrics" in mwb.sheetnames else mwb.active
        if mws is None:
            continue
        mrows = list(mws.iter_rows(values_only=True))
        if len(mrows) < 2:
            continue
        mheaders = list(mrows[0])
        col_idx = {m: mheaders.index(m) for m in metric_cols if m in mheaders}
        sums = {m: [] for m in col_idx}
        for r in mrows[1:]:
            if all(c is None or c == "" for c in r):
                continue
            for m, idx in col_idx.items():
                if idx < len(r) and isinstance(r[idx], (int, float)):
                    sums[m].append(float(r[idx]))
        avgs = {m: (sum(v)/len(v) if v else None) for m, v in sums.items()}
        metrics_data[ver] = {m: v for m, v in avgs.items() if v is not None}
        n = len(mrows) - 1
        all_avg = sum(avgs[m] for m in metric_cols if avgs[m] is not None) / max(
            sum(1 for m in metric_cols if avgs[m] is not None), 1
        )
        metrics_summary.append([
            ver,
            LABELS.get(ver, ver),
            n,
            *[f"{avgs[m]:.3f}" if avgs[m] is not None else "-" for m in metric_cols],
            f"{all_avg:.3f}",
        ])
    for cell in metrics_summary[1]:
        cell.font = bold
        cell.fill = header_fill
    metrics_summary.column_dimensions["A"].width = 12
    metrics_summary.column_dimensions["B"].width = 50
    for c in "CDEFGH":
        metrics_summary.column_dimensions[c].width = 18

    # Δ vs v1
    metrics_summary.append([])
    metrics_summary.append(["v1 대비 Δ"])
    for ver, data in metrics_data.items():
        if ver == "v1":
            continue
        v1_data = metrics_data.get("v1", {})
        delta_cells = []
        for m in metric_cols:
            if m in data and m in v1_data:
                d = data[m] - v1_data[m]
                delta_cells.append(f"{d:+.3f}")
            else:
                delta_cells.append("-")
        metrics_summary.append([ver, LABELS.get(ver, ver), "", *delta_cells, ""])

    # 평가 환경 메모
    metrics_summary.append([])
    metrics_summary.append([
        "측정 환경",
        "gemini-3.1-flash-lite-preview, langchain 우회 (Gemini SDK 직접), concurrency=5, 봇당 ~45초",
    ])
    metrics_summary.append([
        "RAGAS hang 원인",
        "gemini-2.5-pro RPD 1,000 한도 초과 — flash-lite-preview RPD 150K 사용으로 해결",
    ])

    # === 핵심 결론 시트 ===
    concl = wb.create_sheet("핵심 결론")
    concl.append(["항목", "결과"])
    concl.append(["옵션 B (prefix) — NotebookLM", "v1 0.550 vs v2 0.540 → -1%p (노이즈 범위 ±2~3%p, 통계 미달 가능)"])
    concl.append(["옵션 F 재측정 (v3new) — NotebookLM", "v1 0.550 vs v3new 0.540 → -1%p (옛 v3prodOld -2%p의 절반은 봇 설정 차이)"])
    concl.append(["옵션 F 동기화 (v3prodSync) — NotebookLM", "v1 0.550 vs v3prodSync 0.560 → +1%p (LLM 비결정성으로 v3new와 ±2%p 차이)"])
    concl.append(["옵션 F PoC (v3poc) — NotebookLM", "+6%p (0.550 → 0.610) ❌ 봇 설정 불공평 + 1권만 paragraph 시너지 — 신뢰도 낮음"])
    concl.append([])
    concl.append(["★ 종합", "v1/v2/v3 모두 ±2%p 노이즈 범위 — 실용적 동등. v1(sentence) 그대로 유지가 안전"])
    concl.append([])
    concl.append(["4메트릭 직접 측정 (n=100)", "v3prodSync 모든 메트릭 1위 — faith 0.887, ctx_prec 0.754, ctx_recall 0.817, ans_rel 0.927"])
    concl.append(["v1 vs v3prodSync Δ", "+0.018 / +0.028 / +0.017 / +0.045 — 4메트릭 모두 v3 약간 우월 (paragraph)"])
    concl.append([])
    concl.append(["NotebookLM hit rate 자체 노이즈", "같은 데이터/봇으로 측정해도 ±2%p 변동 (LLM 답변 비결정성)"])
    concl.append(["봇 설정 공평성 핵심", "system_prompt + search_tiers 동일성이 측정 결정변수 — 옵션 B 패턴 ('all' 봇 collection_main 토글) 표준"])
    for cell in concl[1]:
        cell.font = bold
        cell.fill = header_fill
    concl.column_dimensions["A"].width = 35
    concl.column_dimensions["B"].width = 100

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"\n저장: {args.output}")
    print("  - 시트1: A_B 6단계 비교")
    print("  - 시트2: 요약 (6 방식 hit율)")
    print("  - 시트3: 봇 스펙 (공평성 검증) ★")
    print("  - 시트4: 4메트릭 직접 측정 (n=100) ★")
    print("  - 시트5: 핵심 결론")
    return 0


if __name__ == "__main__":
    sys.exit(main())
