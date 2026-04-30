"""Phase 2 새 평가셋 통합 보고서 (A vs F 옵션 비교).

3가지 평가 결과를 한 xlsx로 통합:
  - RAGAS 4메트릭 (A vs F, eval_metrics_direct.py 출력 xlsx 두 개)
  - LLM-Judge 정성 4메트릭 + 키워드 F1 (eval_llm_judge.py 출력 csv 두 개)
  - Codex 독립 검토 텍스트 (선택, 결과 md 파일 한 개)

또한 결론 보고서 md 자동 생성.

사용:
    PYTHONPATH=. uv run python scripts/build_phase2_combined_report.py \\
        --ragas-a ~/Downloads/ragas_A_new50_*.xlsx \\
        --ragas-f ~/Downloads/ragas_F_new50_*.xlsx \\
        --judge-a-csv ~/Downloads/llm_judge_A_new50_detail.csv \\
        --judge-f-csv ~/Downloads/llm_judge_F_new50_detail.csv \\
        --codex-md ~/Downloads/codex_review_new_dataset.md \\
        --output ~/Downloads/ab_comparison_new_dataset.xlsx \\
        --report ~/Downloads/phase2_new_dataset_report.md
"""
from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path
from statistics import mean

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


RAGAS_METRICS = ["faithfulness", "context_precision", "context_recall", "response_relevancy"]
JUDGE_METRICS = ["answer_correctness", "context_relevance", "context_faithfulness", "context_recall"]


def load_ragas(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["metrics"] if "metrics" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        d = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
        out.append(d)
    return out


def load_judge_csv(csv_path: Path) -> list[dict]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def avg_metric(rows: list[dict], key: str) -> float | None:
    vals = []
    for r in rows:
        v = r.get(key)
        if v is None or v == "":
            continue
        try:
            vals.append(float(v))
        except (ValueError, TypeError):
            continue
    return mean(vals) if vals else None


def by_level(rows: list[dict]) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {}
    for r in rows:
        L = str(r.get("level", "")).strip()
        for x in ("L1", "L2", "L3", "L4", "L5"):
            if x in L:
                out.setdefault(x, []).append(r)
                break
    return out


def write_ragas_sheet(ws, a_rows: list[dict], f_rows: list[dict]) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    diff_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.append(["RAGAS 4메트릭 — 새 평가셋 50문항 (A: sentence vs F: paragraph)"])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=5)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    # 전체 평균
    ws.append(["메트릭", "A 평균", "F 평균", "차이 (F-A)", "n_A / n_F"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    for m in RAGAS_METRICS:
        a_v = avg_metric(a_rows, m)
        f_v = avg_metric(f_rows, m)
        diff = (f_v - a_v) if (a_v is not None and f_v is not None) else None
        n_a = sum(1 for r in a_rows if r.get(m) not in (None, ""))
        n_f = sum(1 for r in f_rows if r.get(m) not in (None, ""))
        row = [
            m,
            f"{a_v:.3f}" if a_v is not None else "-",
            f"{f_v:.3f}" if f_v is not None else "-",
            f"{diff:+.3f}" if diff is not None else "-",
            f"{n_a} / {n_f}",
        ]
        ws.append(row)
        if diff is not None:
            ws.cell(row=ws.max_row, column=4).fill = diff_fill

    # 평균 종합
    a_total = mean([avg_metric(a_rows, m) or 0 for m in RAGAS_METRICS])
    f_total = mean([avg_metric(f_rows, m) or 0 for m in RAGAS_METRICS])
    ws.append(["**4메트릭 평균**", f"{a_total:.3f}", f"{f_total:.3f}", f"{f_total-a_total:+.3f}", "-"])
    for c in ws[ws.max_row]:
        c.font = bold

    # L별
    ws.append([])
    ws.append(["L별 평균 (A 평균 / F 평균)"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["난이도", "n", "faithfulness", "context_precision", "context_recall", "response_relevancy"])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = header_fill

    a_by_L = by_level(a_rows)
    f_by_L = by_level(f_rows)
    for L in ("L1", "L2", "L3", "L4", "L5"):
        a_g = a_by_L.get(L, [])
        f_g = f_by_L.get(L, [])
        cells = [L, f"{len(a_g)}/{len(f_g)}"]
        for m in RAGAS_METRICS:
            a_v = avg_metric(a_g, m)
            f_v = avg_metric(f_g, m)
            cells.append(f"{a_v:.2f} / {f_v:.2f}" if (a_v is not None and f_v is not None) else "-")
        ws.append(cells)


def write_judge_sheet(ws, a_rows: list[dict], f_rows: list[dict]) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    diff_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.append(["LLM-Judge 정성 4메트릭 — 새 평가셋 50문항"])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=5)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["메트릭 (1~5)", "A 평균", "F 평균", "차이 (F-A)", "n_A / n_F"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    for m in JUDGE_METRICS:
        a_v = avg_metric(a_rows, m)
        f_v = avg_metric(f_rows, m)
        diff = (f_v - a_v) if (a_v is not None and f_v is not None) else None
        n_a = sum(1 for r in a_rows if r.get(m) not in (None, ""))
        n_f = sum(1 for r in f_rows if r.get(m) not in (None, ""))
        row = [m, f"{a_v:.2f}" if a_v else "-", f"{f_v:.2f}" if f_v else "-",
               f"{diff:+.2f}" if diff is not None else "-", f"{n_a} / {n_f}"]
        ws.append(row)
        if diff is not None:
            ws.cell(row=ws.max_row, column=4).fill = diff_fill

    # 총점
    a_total = avg_metric(a_rows, "total_score")
    f_total = avg_metric(f_rows, "total_score")
    diff_t = (f_total - a_total) if (a_total is not None and f_total is not None) else None
    ws.append([
        "**총점 (4~20)**",
        f"{a_total:.2f}" if a_total else "-",
        f"{f_total:.2f}" if f_total else "-",
        f"{diff_t:+.2f}" if diff_t is not None else "-",
        "-",
    ])
    for c in ws[ws.max_row]:
        c.font = bold

    # 등급
    def grade(t):
        if t is None: return "-"
        if t >= 18: return "우수"
        if t >= 15: return "양호"
        if t >= 12: return "보통"
        return "미흡"
    ws.append([])
    ws.append(["등급", grade(a_total), grade(f_total)])
    ws.cell(row=ws.max_row, column=1).font = bold


def write_keyword_f1_sheet(ws, a_rows: list[dict], f_rows: list[dict]) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    ws.append(["키워드 포함률 F1 — 새 평가셋 50문항"])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    a_f1 = avg_metric(a_rows, "kw_f1")
    f_f1 = avg_metric(f_rows, "kw_f1")
    diff = (f_f1 - a_f1) if (a_f1 is not None and f_f1 is not None) else None

    ws.append(["지표", "A 평균", "F 평균", "차이 (F-A)"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    ws.append(["kw_f1 (참고키워드 포함률)",
               f"{a_f1:.3f}" if a_f1 else "-",
               f"{f_f1:.3f}" if f_f1 else "-",
               f"{diff:+.3f}" if diff is not None else "-"])

    # L별
    ws.append([])
    ws.append(["L별 F1 평균 (A / F)"])
    ws.cell(row=ws.max_row, column=1).font = bold

    a_by_L = by_level(a_rows)
    f_by_L = by_level(f_rows)
    ws.append(["난이도", "n", "kw_f1"])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = header_fill
    for L in ("L1", "L2", "L3", "L4", "L5"):
        a_g = a_by_L.get(L, [])
        f_g = f_by_L.get(L, [])
        a_v = avg_metric(a_g, "kw_f1")
        f_v = avg_metric(f_g, "kw_f1")
        ws.append([L, f"{len(a_g)}/{len(f_g)}",
                   f"{a_v:.2f} / {f_v:.2f}" if (a_v is not None and f_v is not None) else "-"])


def write_codex_sheet(ws, codex_text: str) -> None:
    ws.append(["Codex 독립 검토"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for line in codex_text.split("\n"):
        ws.append([line])


def write_legacy_sheet(ws) -> None:
    """이전 100선 결과 핵심 (handoff 메모리 기반)."""
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    ws.append(["[참고] 이전 100선 측정 결과 (gemini-3.1-flash-lite-preview, RAGAS 4메트릭)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["옵션", "faithfulness", "context_precision", "context_recall", "response_relevancy", "평균"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    ws.append(["A (v1, sentence)", "0.869", "0.726", "0.800", "0.882", "0.819"])
    ws.append(["B (v2, prefix)", "0.864", "0.746", "0.784", "0.890", "0.821"])
    ws.append(["F (v3prodSync)", "0.887", "0.754", "0.817", "0.927", "0.847"])
    ws.append([])
    ws.append(["주의: 동일 100문항 A→B→F 순차 측정 → semantic_cache 영향 가능성 의심됨."])
    ws.append(["새 평가셋 50문항으로 캐시 영향 격리 + 측정 순서 반전(F 먼저) 재검증."])


def build_report_md(args, a_ragas, f_ragas, a_judge, f_judge, codex_text: str | None) -> str:
    a_total_ragas = mean([avg_metric(a_ragas, m) or 0 for m in RAGAS_METRICS])
    f_total_ragas = mean([avg_metric(f_ragas, m) or 0 for m in RAGAS_METRICS])
    a_total_judge = avg_metric(a_judge, "total_score")
    f_total_judge = avg_metric(f_judge, "total_score")
    a_f1 = avg_metric(a_judge, "kw_f1")
    f_f1 = avg_metric(f_judge, "kw_f1")

    def winner(av, fv, label):
        if av is None or fv is None: return "?"
        d = fv - av
        if abs(d) < 0.01: return f"{label}: 무차이"
        return f"{label}: {'F' if d > 0 else 'A'} 우월 ({d:+.3f})"

    lines = []
    lines.append("# Phase 2 — 새 평가셋 50문항 A vs F 재검증")
    lines.append("")
    lines.append(f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")
    lines.append("## 종합 결론")
    lines.append("")
    lines.append(f"- RAGAS 4메트릭 평균 — A: **{a_total_ragas:.3f}**, F: **{f_total_ragas:.3f}** ({winner(a_total_ragas, f_total_ragas, 'F-A')})")
    if a_total_judge is not None:
        lines.append(f"- LLM-Judge 총점 (4~20) — A: **{a_total_judge:.2f}**, F: **{f_total_judge:.2f}** ({winner(a_total_judge, f_total_judge, 'F-A')})")
    if a_f1 is not None:
        lines.append(f"- 키워드 F1 — A: **{a_f1:.3f}**, F: **{f_f1:.3f}** ({winner(a_f1, f_f1, 'F-A')})")
    lines.append("")
    lines.append("## RAGAS 메트릭별 비교")
    lines.append("")
    lines.append("| 메트릭 | A 평균 | F 평균 | 차이 (F-A) |")
    lines.append("|---|---:|---:|---:|")
    for m in RAGAS_METRICS:
        a_v = avg_metric(a_ragas, m)
        f_v = avg_metric(f_ragas, m)
        d = (f_v - a_v) if (a_v is not None and f_v is not None) else None
        lines.append(f"| {m} | {a_v:.3f} | {f_v:.3f} | {d:+.3f} |" if d is not None else f"| {m} | - | - | - |")
    lines.append("")
    if a_judge:
        lines.append("## LLM-Judge 메트릭별 비교")
        lines.append("")
        lines.append("| 메트릭 | A 평균 | F 평균 | 차이 |")
        lines.append("|---|---:|---:|---:|")
        for m in JUDGE_METRICS:
            a_v = avg_metric(a_judge, m)
            f_v = avg_metric(f_judge, m)
            d = (f_v - a_v) if (a_v is not None and f_v is not None) else None
            lines.append(f"| {m} | {a_v:.2f} | {f_v:.2f} | {d:+.2f} |" if d is not None else f"| {m} | - | - | - |")
        lines.append("")
    lines.append("## 측정 조건")
    lines.append("")
    lines.append("- 평가셋: 참부모님 생애와 통일원리 문답 학습서 (50문항, ID 101~150)")
    lines.append("- L분포: L1~L5 각 10건씩 균등")
    lines.append("- 측정 순서: F (paragraph) 먼저 → A (sentence) — 캐시 영향 격리")
    lines.append("- 캐시: 측정 시작 시 `ensure_cache_collection`으로 빈 컬렉션 보장")
    lines.append("- 평가 모델: gemini-3.1-flash-lite-preview, temperature=0")
    lines.append("")
    if codex_text:
        lines.append("## Codex 독립 검토 (10건 stratified)")
        lines.append("")
        lines.append(codex_text[:3000])
        if len(codex_text) > 3000:
            lines.append("")
            lines.append(f"... (전체 {len(codex_text)}자, 별도 md 참고)")
        lines.append("")
    lines.append("## 이전 100선 결과 (참고)")
    lines.append("")
    lines.append("| 옵션 | faithfulness | context_precision | context_recall | response_relevancy | 평균 |")
    lines.append("|---|---:|---:|---:|---:|---:|")
    lines.append("| A (v1) | 0.869 | 0.726 | 0.800 | 0.882 | 0.819 |")
    lines.append("| B (v2) | 0.864 | 0.746 | 0.784 | 0.890 | 0.821 |")
    lines.append("| F (v3) | 0.887 | 0.754 | 0.817 | 0.927 | **0.847** |")
    lines.append("")
    lines.append("> 동일 100문항 A→B→F 순차 측정으로 cache 영향 의심됨. 본 50문항 재검증으로 확인.")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ragas-a", required=True, type=Path)
    parser.add_argument("--ragas-f", required=True, type=Path)
    parser.add_argument("--judge-a-csv", required=True, type=Path)
    parser.add_argument("--judge-f-csv", required=True, type=Path)
    parser.add_argument("--codex-md", type=Path, default=None)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    a_ragas = load_ragas(args.ragas_a)
    f_ragas = load_ragas(args.ragas_f)
    a_judge = load_judge_csv(args.judge_a_csv)
    f_judge = load_judge_csv(args.judge_f_csv)
    codex_text = args.codex_md.read_text(encoding="utf-8") if (args.codex_md and args.codex_md.exists()) else None

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
    write_ragas_sheet(wb.create_sheet("RAGAS 4메트릭"), a_ragas, f_ragas)
    write_judge_sheet(wb.create_sheet("LLM-Judge 정성"), a_judge, f_judge)
    write_keyword_f1_sheet(wb.create_sheet("키워드 F1"), a_judge, f_judge)
    if codex_text:
        write_codex_sheet(wb.create_sheet("Codex 검토"), codex_text)
    write_legacy_sheet(wb.create_sheet("이전 100선 (참고)"))

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    report = build_report_md(args, a_ragas, f_ragas, a_judge, f_judge, codex_text)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(report, encoding="utf-8")

    print(f"통합 xlsx: {args.output}")
    print(f"결론 보고서: {args.report}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
