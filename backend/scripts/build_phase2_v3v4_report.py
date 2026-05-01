"""Phase 2 v3 vs v4 통합 보고서 (paragraph vs paragraph+metadata-prefix).

build_phase2_combined_report.py의 v3/v4 라벨 + 평가셋 분리 분석 변형.

분리 분석:
  100문항 합산 결과 + 평가셋별 분리 (퀴즈DB 50 vs 학습서 50)
  - seed JSON id 패턴이 "{label}_{idx:03d}" 형식이라 ID로는 분리 어려움
  - source_file 필드로 분리 (build_ragas_seeds_from_ab.py가 source_file에 xlsx 파일명 저장)
  - LLM-Judge CSV에는 source_file 없음 → seed JSON과 join 필요
  - 단순화: 시드 JSON id 순서대로 0~49는 첫 파일, 50~99는 두 번째 파일로 가정 (eval_notebooklm_qa의 --light/--level5 합치기 순서)

사용:
    PYTHONPATH=. uv run python scripts/build_phase2_v3v4_report.py \\
        --ragas-v3 ~/Downloads/ragas_v3_n100_*.xlsx \\
        --ragas-v4 ~/Downloads/ragas_v4_n100_*.xlsx \\
        --judge-v3-csv ~/Downloads/llm_judge_v3_n100_*_detail.csv \\
        --judge-v4-csv ~/Downloads/llm_judge_v4_n100_*_detail.csv \\
        --codex-md ~/Downloads/codex_review_v3_v4.md \\
        --output ~/Downloads/v3_v4_comparison.xlsx \\
        --report ~/Downloads/phase2_v4_PoC_report.md
"""
from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path
from statistics import mean

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


RAGAS_METRICS = ["faithfulness", "context_precision", "context_recall", "response_relevancy"]
JUDGE_METRICS = ["answer_correctness", "context_relevance", "context_faithfulness", "context_recall"]


def load_ragas(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["metrics"] if "metrics" in wb.sheetnames else wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        out.append({headers[i]: r[i] if i < len(r) else None for i in range(len(headers))})
    return out


def load_judge_csv(csv_path: Path) -> list[dict]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def avg_metric(rows: list[dict], key: str) -> float | None:
    vals = []
    for r in rows:
        v = r.get(key)
        if v is None or v == "":
            continue
        try:
            vals.append(float(v))
        except (ValueError, TypeError):
            continue
    return mean(vals) if vals else None


def by_level(rows: list[dict]) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {}
    for r in rows:
        L = str(r.get("level", "")).strip()
        for x in ("L1", "L2", "L3", "L4", "L5"):
            if x in L:
                out.setdefault(x, []).append(r)
                break
    return out


def split_by_dataset(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """id 인덱스로 분리: idx<=50 → 퀴즈DB, idx>50 → 학습서.

    eval_notebooklm_qa.py --light(퀴즈DB) --level5(학습서) 합치기 순서 가정.
    seed JSON id 형식: "{label}_{idx:03d}"  (1-indexed)
    """
    quiz, study = [], []
    for r in rows:
        rid = str(r.get("id", ""))
        # id에서 마지막 3자리 추출
        parts = rid.split("_")
        if not parts:
            continue
        try:
            idx = int(parts[-1])
        except ValueError:
            continue
        if idx <= 50:
            quiz.append(r)
        else:
            study.append(r)
    return quiz, study


def winner(v3: float | None, v4: float | None, label_v3="v3", label_v4="v4") -> str:
    if v3 is None or v4 is None:
        return "?"
    d = v4 - v3
    if abs(d) < 0.005:
        return f"동등"
    return f"{label_v4 if d > 0 else label_v3} ({d:+.3f})"


def write_ragas_sheet(ws, v3, v4) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    diff_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.append(["RAGAS 4메트릭 — 새 평가셋 100문항 (v3: paragraph vs v4: paragraph+meta-prefix)"])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=5)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["메트릭", "v3 평균", "v4 평균", "차이 (v4-v3)", "n_v3 / n_v4"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    for m in RAGAS_METRICS:
        v3_v = avg_metric(v3, m)
        v4_v = avg_metric(v4, m)
        diff = (v4_v - v3_v) if (v3_v is not None and v4_v is not None) else None
        n_v3 = sum(1 for r in v3 if r.get(m) not in (None, ""))
        n_v4 = sum(1 for r in v4 if r.get(m) not in (None, ""))
        ws.append([
            m,
            f"{v3_v:.3f}" if v3_v is not None else "-",
            f"{v4_v:.3f}" if v4_v is not None else "-",
            f"{diff:+.3f}" if diff is not None else "-",
            f"{n_v3} / {n_v4}",
        ])
        if diff is not None:
            ws.cell(row=ws.max_row, column=4).fill = diff_fill

    v3_total = mean([avg_metric(v3, m) or 0 for m in RAGAS_METRICS])
    v4_total = mean([avg_metric(v4, m) or 0 for m in RAGAS_METRICS])
    ws.append(["**4메트릭 평균**", f"{v3_total:.3f}", f"{v4_total:.3f}", f"{v4_total-v3_total:+.3f}", "-"])
    for c in ws[ws.max_row]:
        c.font = bold

    # 평가셋별 분리
    v3_quiz, v3_study = split_by_dataset(v3)
    v4_quiz, v4_study = split_by_dataset(v4)
    ws.append([])
    ws.append(["평가셋 분리 분석"])
    ws.cell(row=ws.max_row, column=1).font = bold
    for ds_label, v3_ds, v4_ds in [("퀴즈DB (idx 1~50)", v3_quiz, v4_quiz), ("학습서 (idx 51~100)", v3_study, v4_study)]:
        ws.append([ds_label])
        ws.cell(row=ws.max_row, column=1).font = bold
        for m in RAGAS_METRICS:
            v3_v = avg_metric(v3_ds, m)
            v4_v = avg_metric(v4_ds, m)
            ws.append([
                f"  {m}",
                f"{v3_v:.3f}" if v3_v else "-",
                f"{v4_v:.3f}" if v4_v else "-",
                f"{v4_v-v3_v:+.3f}" if (v3_v and v4_v) else "-",
                f"{len(v3_ds)} / {len(v4_ds)}",
            ])

    # L별
    ws.append([])
    ws.append(["L별 평균 (v3 / v4)"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["난이도", "n", "faithfulness", "context_precision", "context_recall", "response_relevancy"])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = header_fill

    v3_by_L = by_level(v3)
    v4_by_L = by_level(v4)
    for L in ("L1", "L2", "L3", "L4", "L5"):
        v3_g = v3_by_L.get(L, [])
        v4_g = v4_by_L.get(L, [])
        cells = [L, f"{len(v3_g)}/{len(v4_g)}"]
        for m in RAGAS_METRICS:
            v3_v = avg_metric(v3_g, m)
            v4_v = avg_metric(v4_g, m)
            cells.append(f"{v3_v:.2f} / {v4_v:.2f}" if (v3_v is not None and v4_v is not None) else "-")
        ws.append(cells)


def write_judge_sheet(ws, v3, v4) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    diff_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.append(["LLM-Judge 정성 4메트릭 — 100문항 v3 vs v4"])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=5)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["메트릭 (1~5)", "v3 평균", "v4 평균", "차이", "n_v3 / n_v4"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    for m in JUDGE_METRICS:
        v3_v = avg_metric(v3, m)
        v4_v = avg_metric(v4, m)
        diff = (v4_v - v3_v) if (v3_v is not None and v4_v is not None) else None
        n_v3 = sum(1 for r in v3 if r.get(m) not in (None, ""))
        n_v4 = sum(1 for r in v4 if r.get(m) not in (None, ""))
        ws.append([m,
                   f"{v3_v:.2f}" if v3_v else "-",
                   f"{v4_v:.2f}" if v4_v else "-",
                   f"{diff:+.2f}" if diff is not None else "-",
                   f"{n_v3} / {n_v4}"])
        if diff is not None:
            ws.cell(row=ws.max_row, column=4).fill = diff_fill

    v3_total = avg_metric(v3, "total_score")
    v4_total = avg_metric(v4, "total_score")
    diff_t = (v4_total - v3_total) if (v3_total and v4_total) else None
    ws.append(["**총점 (4~20)**",
               f"{v3_total:.2f}" if v3_total else "-",
               f"{v4_total:.2f}" if v4_total else "-",
               f"{diff_t:+.2f}" if diff_t is not None else "-",
               "-"])
    for c in ws[ws.max_row]:
        c.font = bold

    # 평가셋별 분리
    v3_quiz, v3_study = split_by_dataset(v3)
    v4_quiz, v4_study = split_by_dataset(v4)
    ws.append([])
    ws.append(["평가셋 분리 분석 — 총점 (4~20)"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["평가셋", "v3", "v4", "차이", "n"])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = header_fill
    for ds_label, v3_ds, v4_ds in [("퀴즈DB (idx 1~50)", v3_quiz, v4_quiz), ("학습서 (idx 51~100)", v3_study, v4_study)]:
        v3_v = avg_metric(v3_ds, "total_score")
        v4_v = avg_metric(v4_ds, "total_score")
        ws.append([ds_label,
                   f"{v3_v:.2f}" if v3_v else "-",
                   f"{v4_v:.2f}" if v4_v else "-",
                   f"{v4_v-v3_v:+.2f}" if (v3_v and v4_v) else "-",
                   f"{len(v3_ds)} / {len(v4_ds)}"])

    # L별
    ws.append([])
    ws.append(["L별 총점 (v3 / v4) — 채택 기준 v4 L2 ≥ 11.5 + RAGAS 평균 v4 ≥ v3"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["난이도", "v3 총점", "v4 총점", "차이", "비고"])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = header_fill
    v3_by_L = by_level(v3)
    v4_by_L = by_level(v4)
    for L in ("L1", "L2", "L3", "L4", "L5"):
        v3_g = v3_by_L.get(L, [])
        v4_g = v4_by_L.get(L, [])
        v3_v = avg_metric(v3_g, "total_score")
        v4_v = avg_metric(v4_g, "total_score")
        note = ""
        if L == "L2" and v4_v is not None:
            note = "✅ 채택 기준 통과 (≥11.5)" if v4_v >= 11.5 else "❌ 채택 기준 미달 (<11.5)"
        ws.append([L,
                   f"{v3_v:.2f}" if v3_v else "-",
                   f"{v4_v:.2f}" if v4_v else "-",
                   f"{v4_v-v3_v:+.2f}" if (v3_v and v4_v) else "-",
                   note])


def write_keyword_f1_sheet(ws, v3, v4) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    ws.append(["키워드 포함률 F1 — 100문항 v3 vs v4"])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=4)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    v3_f1 = avg_metric(v3, "kw_f1")
    v4_f1 = avg_metric(v4, "kw_f1")
    ws.append(["지표", "v3 평균", "v4 평균", "차이"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    ws.append(["kw_f1 (참고키워드 포함률)",
               f"{v3_f1:.3f}" if v3_f1 else "-",
               f"{v4_f1:.3f}" if v4_f1 else "-",
               f"{v4_f1-v3_f1:+.3f}" if (v3_f1 and v4_f1) else "-"])


def write_codex_sheet(ws, codex_text: str) -> None:
    ws.append(["Codex v3 vs v4 독립 검토"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for line in codex_text.split("\n"):
        ws.append([line])


def write_decision_sheet(ws, v3_ragas, v4_ragas, v3_judge, v4_judge) -> None:
    bold = Font(bold=True)

    v3_total_ragas = mean([avg_metric(v3_ragas, m) or 0 for m in RAGAS_METRICS])
    v4_total_ragas = mean([avg_metric(v4_ragas, m) or 0 for m in RAGAS_METRICS])
    v3_l2 = avg_metric(by_level(v3_judge).get("L2", []), "total_score")
    v4_l2 = avg_metric(by_level(v4_judge).get("L2", []), "total_score")

    cond1 = v4_total_ragas >= v3_total_ragas
    cond2 = (v4_l2 is not None) and v4_l2 >= 11.5

    ws.append(["v4 채택 결정 (사용자 확정 기준)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["조건", "기준", "v3 값", "v4 값", "통과"])
    ws.append([
        "조건 1: RAGAS 4메트릭 평균",
        "v4 ≥ v3",
        f"{v3_total_ragas:.3f}",
        f"{v4_total_ragas:.3f}",
        "✅" if cond1 else "❌",
    ])
    ws.append([
        "조건 2: LLM-Judge L2 총점",
        "v4 ≥ 11.5",
        f"{v3_l2:.2f}" if v3_l2 else "-",
        f"{v4_l2:.2f}" if v4_l2 else "-",
        "✅" if cond2 else "❌",
    ])
    ws.append([])
    final = "v4 채택" if (cond1 and cond2) else "v3 유지 (L2 보강은 별도 plan)"
    ws.append(["최종 판정", final])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True, size=12)


def build_report_md(v3_ragas, v4_ragas, v3_judge, v4_judge, codex_text: str | None) -> str:
    v3_total_ragas = mean([avg_metric(v3_ragas, m) or 0 for m in RAGAS_METRICS])
    v4_total_ragas = mean([avg_metric(v4_ragas, m) or 0 for m in RAGAS_METRICS])
    v3_total_judge = avg_metric(v3_judge, "total_score")
    v4_total_judge = avg_metric(v4_judge, "total_score")
    v3_l2 = avg_metric(by_level(v3_judge).get("L2", []), "total_score")
    v4_l2 = avg_metric(by_level(v4_judge).get("L2", []), "total_score")

    cond1 = v4_total_ragas >= v3_total_ragas
    cond2 = (v4_l2 is not None) and v4_l2 >= 11.5

    lines = []
    lines.append("# Phase 2 v4 PoC — paragraph + metadata prefix injection")
    lines.append("")
    lines.append(f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")
    lines.append("## 종합 결론")
    lines.append("")
    final = "**v4 채택**" if (cond1 and cond2) else "**v3 유지** (L2 보강은 별도 plan)"
    lines.append(f"### 최종 판정: {final}")
    lines.append("")
    lines.append(f"- 조건 1 (RAGAS v4 ≥ v3): {'✅ 통과' if cond1 else '❌ 미달'} — v3: {v3_total_ragas:.3f} / v4: {v4_total_ragas:.3f}")
    lines.append(f"- 조건 2 (LLM-Judge L2 ≥ 11.5): {'✅ 통과' if cond2 else '❌ 미달'} — v3: {v3_l2:.2f} / v4: {v4_l2:.2f}" if v3_l2 else "- 조건 2: 측정 미완")
    lines.append("")
    lines.append("## RAGAS 4메트릭")
    lines.append("")
    lines.append("| 메트릭 | v3 | v4 | 차이 |")
    lines.append("|---|---:|---:|---:|")
    for m in RAGAS_METRICS:
        v3_v = avg_metric(v3_ragas, m)
        v4_v = avg_metric(v4_ragas, m)
        d = (v4_v - v3_v) if (v3_v and v4_v) else None
        lines.append(f"| {m} | {v3_v:.3f} | {v4_v:.3f} | {d:+.3f} |" if d is not None else f"| {m} | - | - | - |")
    lines.append(f"| **평균** | **{v3_total_ragas:.3f}** | **{v4_total_ragas:.3f}** | **{v4_total_ragas-v3_total_ragas:+.3f}** |")
    lines.append("")
    lines.append("## LLM-Judge 정성 4메트릭")
    lines.append("")
    lines.append("| 메트릭 | v3 | v4 | 차이 |")
    lines.append("|---|---:|---:|---:|")
    for m in JUDGE_METRICS:
        v3_v = avg_metric(v3_judge, m)
        v4_v = avg_metric(v4_judge, m)
        d = (v4_v - v3_v) if (v3_v and v4_v) else None
        lines.append(f"| {m} | {v3_v:.2f} | {v4_v:.2f} | {d:+.2f} |" if d is not None else f"| {m} | - | - | - |")
    if v3_total_judge:
        lines.append(f"| **총점** | **{v3_total_judge:.2f}** | **{v4_total_judge:.2f}** | **{v4_total_judge-v3_total_judge:+.2f}** |")
    lines.append("")
    lines.append("## L별 LLM-Judge 총점")
    lines.append("")
    lines.append("| 난이도 | v3 | v4 | 차이 |")
    lines.append("|---|---:|---:|---:|")
    v3_by_L = by_level(v3_judge)
    v4_by_L = by_level(v4_judge)
    for L in ("L1", "L2", "L3", "L4", "L5"):
        v3_v = avg_metric(v3_by_L.get(L, []), "total_score")
        v4_v = avg_metric(v4_by_L.get(L, []), "total_score")
        d = (v4_v - v3_v) if (v3_v and v4_v) else None
        marker = " 🎯" if L == "L2" else ""
        lines.append(f"| {L}{marker} | {v3_v:.2f} | {v4_v:.2f} | {d:+.2f} |" if d is not None else f"| {L} | - | - | - |")
    lines.append("")
    lines.append("## 평가셋별 분리 분석")
    lines.append("")
    v3_quiz, v3_study = split_by_dataset(v3_judge)
    v4_quiz, v4_study = split_by_dataset(v4_judge)
    lines.append("| 평가셋 | n | v3 총점 | v4 총점 | 차이 |")
    lines.append("|---|---:|---:|---:|---:|")
    for label, v3_ds, v4_ds in [("퀴즈DB", v3_quiz, v4_quiz), ("학습서", v3_study, v4_study)]:
        v3_v = avg_metric(v3_ds, "total_score")
        v4_v = avg_metric(v4_ds, "total_score")
        d = (v4_v - v3_v) if (v3_v and v4_v) else None
        lines.append(f"| {label} | {len(v3_ds)}/{len(v4_ds)} | {v3_v:.2f} | {v4_v:.2f} | {d:+.2f} |" if d is not None else f"| {label} | - | - | - | - |")
    lines.append("")
    lines.append("## 측정 조건")
    lines.append("")
    lines.append("- 평가셋: 통일교 원리 및 섭리 퀴즈 데이터베이스 (50) + 참부모님 생애와 통일원리 문답 학습서 (50) = 100문항")
    lines.append("- 측정 순서: v4 먼저 → v3 (cache 비우기 + ensure 사이)")
    lines.append("- v4 prefix 형식: `[volume / date]` (date 누락 시 `[volume]`)")
    lines.append("- 챗봇 토글: 'all' 봇 collection_main만 변경")
    lines.append("- 평가 모델: gemini-3.1-flash-lite-preview, temperature=0")
    lines.append("- Codex: OpenAI gpt-5-codex (consult mode)")
    lines.append("")
    if codex_text:
        lines.append("## Codex v3 vs v4 독립 검토 (10건 stratified)")
        lines.append("")
        lines.append(codex_text[:5000])
        if len(codex_text) > 5000:
            lines.append("")
            lines.append(f"... (전체 {len(codex_text)}자, 별도 md 참고)")
        lines.append("")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ragas-v3", required=True, type=Path)
    parser.add_argument("--ragas-v4", required=True, type=Path)
    parser.add_argument("--judge-v3-csv", required=True, type=Path)
    parser.add_argument("--judge-v4-csv", required=True, type=Path)
    parser.add_argument("--codex-md", type=Path, default=None)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    v3_ragas = load_ragas(args.ragas_v3)
    v4_ragas = load_ragas(args.ragas_v4)
    v3_judge = load_judge_csv(args.judge_v3_csv)
    v4_judge = load_judge_csv(args.judge_v4_csv)
    codex_text = args.codex_md.read_text(encoding="utf-8") if (args.codex_md and args.codex_md.exists()) else None

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
    write_decision_sheet(wb.create_sheet("결정"), v3_ragas, v4_ragas, v3_judge, v4_judge)
    write_ragas_sheet(wb.create_sheet("RAGAS"), v3_ragas, v4_ragas)
    write_judge_sheet(wb.create_sheet("LLM-Judge"), v3_judge, v4_judge)
    write_keyword_f1_sheet(wb.create_sheet("키워드 F1"), v3_judge, v4_judge)
    if codex_text:
        write_codex_sheet(wb.create_sheet("Codex"), codex_text)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    report = build_report_md(v3_ragas, v4_ragas, v3_judge, v4_judge, codex_text)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(report, encoding="utf-8")

    print(f"v3-v4 통합 xlsx: {args.output}")
    print(f"v4 PoC 결론 보고서: {args.report}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
