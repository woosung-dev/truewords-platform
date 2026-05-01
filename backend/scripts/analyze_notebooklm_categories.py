"""NotebookLM 평가 결과 두 xlsx(베이스라인 vs 액션 적용 후)를 카테고리·사례 단위로 분석.

휴리스틱:
    참고키워드(쉼표/줄바꿈 분리) 토큰 절반 이상이 참고1+참고2+참고3 텍스트에 포함되면 hit.
    토큰 1개뿐이면 그 1개라도 매치되면 hit (보수적 임계).

사용 예:
    PYTHONPATH=. uv run python scripts/analyze_notebooklm_categories.py \
        --baseline ~/Downloads/notebooklm_qa_전체검색봇_평가_튜닝후_20260427_1649.xlsx \
        --treatment ~/Downloads/notebooklm_post_phase1_20260428_light100.xlsx \
        --output ~/Downloads/notebooklm_post_phase1_20260428_category_analysis.xlsx

출력 시트:
    - "카테고리": Level별 baseline/treatment hit율과 delta
    - "회귀": baseline hit, treatment miss인 행
    - "개선": baseline miss, treatment hit인 행
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook


def _load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c) if c is not None else "" for c in rows[0]]
    out: list[dict] = []
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        out.append(dict(zip(header, r)))
    return out


def _cited_text(row: dict) -> str:
    return " ".join(str(row.get(k, "") or "") for k in ("참고1", "참고2", "참고3"))


def _is_hit(row: dict) -> bool:
    keywords = str(row.get("참고 키워드") or row.get("참고키워드") or "")
    tokens = [t.strip() for t in keywords.replace("\n", ",").split(",") if t.strip()]
    if not tokens:
        return False
    cited = _cited_text(row)
    if not cited:
        return False
    matched = sum(1 for t in tokens if t in cited)
    threshold = max(1, len(tokens) // 2)
    return matched >= threshold


def _level(row: dict) -> str:
    raw = str(row.get("Level") or row.get("난이도(Level)") or row.get("질문 수준") or "").strip()
    return raw or "Unknown"


def _category_group(row: dict) -> str:
    return str(row.get("카테고리") or "").strip() or "Unknown"


def _question(row: dict) -> str:
    return str(row.get("질문") or row.get("테스트용 질문") or "")


def _gold_answer(row: dict) -> str:
    return str(row.get("모범답변") or row.get("봇 모범 답변") or "")


def _our_answer(row: dict) -> str:
    for key in ("우리 답변(all)", "우리 답변", "우리답변"):
        v = row.get(key)
        if v:
            return str(v)
    # 헤더 첫 7번째 컬럼이 "우리 답변(<id>)" 형태라 명시 키 없으면 빈 문자열
    return ""


def compute_category_hit_table(baseline: Path, treatment: Path) -> list[dict]:
    base = _load_rows(baseline)
    tune = _load_rows(treatment)
    if not base and not tune:
        return []
    base_by_no = {r.get("번호"): r for r in base}
    tune_by_no = {r.get("번호"): r for r in tune}
    nums = sorted(
        set(base_by_no) | set(tune_by_no),
        key=lambda x: (isinstance(x, str), x),
    )
    levels: dict[str, list[tuple[bool, bool]]] = {}
    for n in nums:
        b = base_by_no.get(n)
        t = tune_by_no.get(n)
        lv = _level(b or t or {})
        levels.setdefault(lv, []).append(
            (_is_hit(b) if b else False, _is_hit(t) if t else False)
        )
    out: list[dict] = []
    for lv in sorted(levels):
        pairs = levels[lv]
        n = len(pairs)
        b_hit = sum(1 for b, _ in pairs if b)
        t_hit = sum(1 for _, t in pairs if t)
        out.append({
            "level": lv,
            "n": n,
            "baseline_hit_rate": b_hit / n if n else 0.0,
            "treatment_hit_rate": t_hit / n if n else 0.0,
            "delta": (t_hit - b_hit) / n if n else 0.0,
        })
    return out


def _diff_rows(baseline: Path, treatment: Path, base_hit: bool, tune_hit: bool) -> list[dict]:
    base = {r.get("번호"): r for r in _load_rows(baseline)}
    tune = {r.get("번호"): r for r in _load_rows(treatment)}
    out: list[dict] = []
    for n in sorted(base.keys() & tune.keys(), key=lambda x: (isinstance(x, str), x)):
        if _is_hit(base[n]) is base_hit and _is_hit(tune[n]) is tune_hit:
            out.append({
                "번호": n,
                "Level": _level(tune[n]),
                "카테고리": _category_group(tune[n]),
                "질문": _question(tune[n]),
                "모범답변": _gold_answer(tune[n]),
                "베이스 답변": _our_answer(base[n]),
                "튜닝 답변": _our_answer(tune[n]),
            })
    return out


def extract_regressions(baseline: Path, treatment: Path) -> list[dict]:
    """베이스라인 hit, treatment miss인 행 (=회귀)."""
    return _diff_rows(baseline, treatment, base_hit=True, tune_hit=False)


def extract_improvements(baseline: Path, treatment: Path) -> list[dict]:
    """베이스라인 miss, treatment hit인 행 (=개선)."""
    return _diff_rows(baseline, treatment, base_hit=False, tune_hit=True)


def _write_excel(
    category: list[dict],
    regressions: list[dict],
    improvements: list[dict],
    output: Path,
) -> None:
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "카테고리"
    ws1.append(["Level", "n", "baseline_hit_rate", "treatment_hit_rate", "delta"])
    for r in category:
        ws1.append([
            r["level"], r["n"],
            round(r["baseline_hit_rate"], 4),
            round(r["treatment_hit_rate"], 4),
            round(r["delta"], 4),
        ])
    ws2 = wb.create_sheet("회귀")
    if regressions:
        ws2.append(list(regressions[0].keys()))
        for r in regressions:
            ws2.append(list(r.values()))
    ws3 = wb.create_sheet("개선")
    if improvements:
        ws3.append(list(improvements[0].keys()))
        for r in improvements:
            ws3.append(list(r.values()))
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--baseline", required=True, type=Path)
    p.add_argument("--treatment", required=True, type=Path)
    p.add_argument("--output", required=True, type=Path)
    args = p.parse_args()
    cat = compute_category_hit_table(args.baseline, args.treatment)
    reg = extract_regressions(args.baseline, args.treatment)
    imp = extract_improvements(args.baseline, args.treatment)
    _write_excel(cat, reg, imp, args.output)
    print(f"분석 완료: {args.output} | 카테고리 {len(cat)}행, 회귀 {len(reg)}건, 개선 {len(imp)}건")
    return 0


if __name__ == "__main__":
    sys.exit(main())
