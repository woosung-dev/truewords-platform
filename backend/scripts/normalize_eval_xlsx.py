"""평가셋 xlsx 컬럼 표준화.

새 평가셋 (예: "통일원리 및 평화사상 단계별 학습 질의서.xlsx")의 컬럼 구조:
  [L (1~5), 명칭, 예상 처리율, 예시 질의, 봇 모범 답변, 참고 키워드]

기존 평가 스크립트(eval_notebooklm_qa.py / build_ragas_seeds_from_ab.py)가 기대하는 형식:
  [번호, 난이도(Level), 카테고리, 테스트용 질문, 봇 모범 답변, 참고 키워드]

변환:
  - col 1 (L=1) + col 2 (명칭) → 난이도(Level) = "L1 (단순 사실 조회)"
  - row 순서 → 번호 (1~100)
  - col 3 (예상 처리율) → 카테고리 (그대로 유지, level이 명확하므로 카테고리 의미 변경 없음)
  - col 4~6 그대로

사용:
    PYTHONPATH=. uv run python scripts/normalize_eval_xlsx.py \\
        --input ~/Downloads/통일원리\\ 및\\ 평화사상\\ 단계별\\ 학습\\ 질의서.xlsx \\
        --output ~/Downloads/통일원리_평화사상_normalized.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


STANDARD_HEADERS = [
    "번호",
    "난이도(Level)",
    "카테고리",
    "테스트용 질문",
    "봇 모범 답변",
    "참고 키워드",
]


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--start-number", type=int, default=1,
                        help="번호 시작값 (다른 평가셋과 ID 충돌 회피용)")
    args = parser.parse_args()

    wb_in = load_workbook(args.input, data_only=True)
    ws_in = wb_in.active
    if ws_in is None:
        print("⚠️  active sheet 없음")
        return 1
    rows = list(ws_in.iter_rows(values_only=True))
    if not rows:
        print("⚠️  비어있음")
        return 1

    headers_in = [str(h) if h is not None else "" for h in rows[0]]
    print(f"입력 헤더: {headers_in}")
    print(f"입력 행 수: {len(rows) - 1}")

    wb_out = Workbook()
    ws_out = wb_out.active
    if ws_out is None:
        return 1
    ws_out.title = "Table 1"
    ws_out.append(STANDARD_HEADERS)
    bold = Font(bold=True)
    for c in ws_out[1]:
        c.font = bold
        c.fill = PatternFill("solid", fgColor="D9E1F2")

    converted = 0
    skipped = 0
    for idx, r in enumerate(rows[1:], 0):
        if not r or all(c is None or c == "" for c in r):
            skipped += 1
            continue
        # col 1: L (1~5 단독 숫자) — 정수 또는 문자열
        L_raw = r[0] if len(r) > 0 else None
        if L_raw is None:
            skipped += 1
            continue
        try:
            L_num = int(str(L_raw).strip())
        except (ValueError, TypeError):
            print(f"  ⚠️  row {idx+1}: L 파싱 실패 ({L_raw!r}), skip")
            skipped += 1
            continue
        if not 1 <= L_num <= 5:
            print(f"  ⚠️  row {idx+1}: L 범위 외 ({L_num}), skip")
            skipped += 1
            continue

        # col 2: 명칭 (단순 사실 조회 등)
        meongching = str(r[1]) if len(r) > 1 and r[1] else ""
        # 표준 형식: "L1 (단순 사실 조회)"
        level_str = f"L{L_num} ({meongching})" if meongching else f"L{L_num}"

        # col 3: 예상 처리율 → 카테고리 슬롯 (정보 유지)
        category = str(r[2]) if len(r) > 2 and r[2] else ""

        # col 4: 예시 질의
        question = str(r[3]) if len(r) > 3 and r[3] else ""
        # col 5: 봇 모범 답변
        gold = str(r[4]) if len(r) > 4 and r[4] else ""
        # col 6: 참고 키워드
        keywords = str(r[5]) if len(r) > 5 and r[5] else ""

        if not question:
            skipped += 1
            continue

        number = args.start_number + converted
        ws_out.append([number, level_str, category, question, gold, keywords])
        converted += 1

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb_out.save(args.output)
    print()
    print(f"=== 변환 완료 ===")
    print(f"  변환: {converted}건")
    print(f"  skip: {skipped}건")
    print(f"  출력: {args.output}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
