"""Phase 2 새 평가셋 3-way 통합 보고서 (A vs B vs F).

3가지 청킹 전략 + 3가지 평가 방식 통합:
  - RAGAS 4메트릭 (A/B/F)
  - LLM-Judge 정성 4메트릭 + 키워드 F1 (A/B/F)
  - Codex 독립 검토 (3-way 비교, 선택)

사용:
    PYTHONPATH=. uv run python scripts/build_phase2_3way_report.py \\
        --ragas-a ~/Downloads/ragas_A_new50_*.xlsx \\
        --ragas-b ~/Downloads/ragas_B_new50_*.xlsx \\
        --ragas-f ~/Downloads/ragas_F_new50_*.xlsx \\
        --judge-a-csv ~/Downloads/llm_judge_A_new50_*_detail.csv \\
        --judge-b-csv ~/Downloads/llm_judge_B_new50_*_detail.csv \\
        --judge-f-csv ~/Downloads/llm_judge_F_new50_*_detail.csv \\
        --codex-md ~/Downloads/codex_review_3way.md \\
        --output ~/Downloads/ab_comparison_3way.xlsx \\
        --report ~/Downloads/phase2_3way_report.md
"""
from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path
from statistics import mean

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


RAGAS_METRICS = ["faithfulness", "context_precision", "context_recall", "response_relevancy"]
JUDGE_METRICS = ["answer_correctness", "context_relevance", "context_faithfulness", "context_recall"]


def load_ragas(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["metrics"] if "metrics" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        out.append({headers[i]: r[i] if i < len(r) else None for i in range(len(headers))})
    return out


def load_judge_csv(csv_path: Path) -> list[dict]:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def avg_metric(rows: list[dict], key: str) -> float | None:
    vals = []
    for r in rows:
        v = r.get(key)
        if v is None or v == "":
            continue
        try:
            vals.append(float(v))
        except (ValueError, TypeError):
            continue
    return mean(vals) if vals else None


def by_level(rows: list[dict]) -> dict[str, list[dict]]:
    out: dict[str, list[dict]] = {}
    for r in rows:
        L = str(r.get("level", "")).strip()
        for x in ("L1", "L2", "L3", "L4", "L5"):
            if x in L:
                out.setdefault(x, []).append(r)
                break
    return out


def winner_3way(a: float | None, b: float | None, f: float | None) -> str:
    """A/B/F 중 최고 옵션 + 격차 표시."""
    vals = {"A": a, "B": b, "F": f}
    valid = {k: v for k, v in vals.items() if v is not None}
    if len(valid) < 2:
        return "?"
    best_k = max(valid, key=lambda k: valid[k])
    sorted_v = sorted(valid.values(), reverse=True)
    gap = sorted_v[0] - sorted_v[1]
    if gap < 0.01:
        return f"{best_k}/동등"
    return f"{best_k} (+{gap:.3f})"


def write_ragas_sheet(ws, a, b, f) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    diff_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.append(["RAGAS 4메트릭 — 새 평가셋 50문항 3-way (A: sentence / B: prefix / F: paragraph)"])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["메트릭", "A 평균", "B 평균", "F 평균", "최고 (격차)", "n_A / n_B / n_F"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    for m in RAGAS_METRICS:
        a_v = avg_metric(a, m)
        b_v = avg_metric(b, m)
        f_v = avg_metric(f, m)
        n_a = sum(1 for r in a if r.get(m) not in (None, ""))
        n_b = sum(1 for r in b if r.get(m) not in (None, ""))
        n_f = sum(1 for r in f if r.get(m) not in (None, ""))
        row = [
            m,
            f"{a_v:.3f}" if a_v is not None else "-",
            f"{b_v:.3f}" if b_v is not None else "-",
            f"{f_v:.3f}" if f_v is not None else "-",
            winner_3way(a_v, b_v, f_v),
            f"{n_a} / {n_b} / {n_f}",
        ]
        ws.append(row)
        ws.cell(row=ws.max_row, column=5).fill = diff_fill

    a_total = mean([avg_metric(a, m) or 0 for m in RAGAS_METRICS])
    b_total = mean([avg_metric(b, m) or 0 for m in RAGAS_METRICS])
    f_total = mean([avg_metric(f, m) or 0 for m in RAGAS_METRICS])
    ws.append([
        "**4메트릭 평균**",
        f"{a_total:.3f}", f"{b_total:.3f}", f"{f_total:.3f}",
        winner_3way(a_total, b_total, f_total),
        "-",
    ])
    for c in ws[ws.max_row]:
        c.font = bold

    ws.append([])
    ws.append(["L별 평균 (A / B / F)"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["난이도", "n", "faithfulness", "context_precision", "context_recall", "response_relevancy"])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = header_fill

    a_by_L = by_level(a)
    b_by_L = by_level(b)
    f_by_L = by_level(f)
    for L in ("L1", "L2", "L3", "L4", "L5"):
        a_g = a_by_L.get(L, [])
        b_g = b_by_L.get(L, [])
        f_g = f_by_L.get(L, [])
        cells = [L, f"{len(a_g)}/{len(b_g)}/{len(f_g)}"]
        for m in RAGAS_METRICS:
            a_v = avg_metric(a_g, m)
            b_v = avg_metric(b_g, m)
            f_v = avg_metric(f_g, m)
            parts = []
            for v in (a_v, b_v, f_v):
                parts.append(f"{v:.2f}" if v is not None else "-")
            cells.append(" / ".join(parts))
        ws.append(cells)


def write_judge_sheet(ws, a, b, f) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    diff_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.append(["LLM-Judge 정성 4메트릭 — 새 평가셋 50문항 3-way"])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(["메트릭 (1~5)", "A 평균", "B 평균", "F 평균", "최고 (격차)", "n_A / n_B / n_F"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    for m in JUDGE_METRICS:
        a_v = avg_metric(a, m)
        b_v = avg_metric(b, m)
        f_v = avg_metric(f, m)
        n_a = sum(1 for r in a if r.get(m) not in (None, ""))
        n_b = sum(1 for r in b if r.get(m) not in (None, ""))
        n_f = sum(1 for r in f if r.get(m) not in (None, ""))
        row = [
            m,
            f"{a_v:.2f}" if a_v else "-",
            f"{b_v:.2f}" if b_v else "-",
            f"{f_v:.2f}" if f_v else "-",
            winner_3way(a_v, b_v, f_v),
            f"{n_a} / {n_b} / {n_f}",
        ]
        ws.append(row)
        ws.cell(row=ws.max_row, column=5).fill = diff_fill

    a_total = avg_metric(a, "total_score")
    b_total = avg_metric(b, "total_score")
    f_total = avg_metric(f, "total_score")
    ws.append([
        "**총점 (4~20)**",
        f"{a_total:.2f}" if a_total else "-",
        f"{b_total:.2f}" if b_total else "-",
        f"{f_total:.2f}" if f_total else "-",
        winner_3way(a_total, b_total, f_total),
        "-",
    ])
    for c in ws[ws.max_row]:
        c.font = bold

    def grade(t):
        if t is None: return "-"
        if t >= 18: return "우수"
        if t >= 15: return "양호"
        if t >= 12: return "보통"
        return "미흡"
    ws.append([])
    ws.append(["등급", grade(a_total), grade(b_total), grade(f_total)])
    ws.cell(row=ws.max_row, column=1).font = bold

    # L별
    ws.append([])
    ws.append(["L별 총점 평균 (A / B / F)"])
    ws.cell(row=ws.max_row, column=1).font = bold
    ws.append(["난이도", "n", "총점 평균"])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = header_fill
    a_by_L = by_level(a)
    b_by_L = by_level(b)
    f_by_L = by_level(f)
    for L in ("L1", "L2", "L3", "L4", "L5"):
        a_g = a_by_L.get(L, [])
        b_g = b_by_L.get(L, [])
        f_g = f_by_L.get(L, [])
        a_v = avg_metric(a_g, "total_score")
        b_v = avg_metric(b_g, "total_score")
        f_v = avg_metric(f_g, "total_score")
        parts = []
        for v in (a_v, b_v, f_v):
            parts.append(f"{v:.2f}" if v is not None else "-")
        ws.append([L, f"{len(a_g)}/{len(b_g)}/{len(f_g)}", " / ".join(parts)])


def write_keyword_f1_sheet(ws, a, b, f) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    ws.append(["키워드 포함률 F1 — 새 평가셋 50문항 3-way"])
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=5)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    a_f1 = avg_metric(a, "kw_f1")
    b_f1 = avg_metric(b, "kw_f1")
    f_f1 = avg_metric(f, "kw_f1")

    ws.append(["지표", "A 평균", "B 평균", "F 평균", "최고 (격차)"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    ws.append([
        "kw_f1 (참고키워드 포함률)",
        f"{a_f1:.3f}" if a_f1 else "-",
        f"{b_f1:.3f}" if b_f1 else "-",
        f"{f_f1:.3f}" if f_f1 else "-",
        winner_3way(a_f1, b_f1, f_f1),
    ])

    ws.append([])
    ws.append(["L별 F1 평균 (A / B / F)"])
    ws.cell(row=ws.max_row, column=1).font = bold

    a_by_L = by_level(a)
    b_by_L = by_level(b)
    f_by_L = by_level(f)
    ws.append(["난이도", "n", "kw_f1"])
    for c in ws[ws.max_row]:
        c.font = bold
        c.fill = header_fill
    for L in ("L1", "L2", "L3", "L4", "L5"):
        a_g = a_by_L.get(L, [])
        b_g = b_by_L.get(L, [])
        f_g = f_by_L.get(L, [])
        a_v = avg_metric(a_g, "kw_f1")
        b_v = avg_metric(b_g, "kw_f1")
        f_v = avg_metric(f_g, "kw_f1")
        parts = []
        for v in (a_v, b_v, f_v):
            parts.append(f"{v:.2f}" if v is not None else "-")
        ws.append([L, f"{len(a_g)}/{len(b_g)}/{len(f_g)}", " / ".join(parts)])


def write_codex_sheet(ws, codex_text: str) -> None:
    ws.append(["Codex 독립 검토 (3-way)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    for line in codex_text.split("\n"):
        ws.append([line])


def write_legacy_sheet(ws) -> None:
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")

    ws.append(["[참고] 이전 100선 측정 결과 (gemini-3.1-flash-lite-preview, RAGAS 4메트릭)"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["옵션", "faithfulness", "context_precision", "context_recall", "response_relevancy", "평균"])
    for c in ws[3]:
        c.font = bold
        c.fill = header_fill
    ws.append(["A (v1, sentence)", "0.869", "0.726", "0.800", "0.882", "0.819"])
    ws.append(["B (v2, prefix)", "0.864", "0.746", "0.784", "0.890", "0.821"])
    ws.append(["F (v3, paragraph)", "0.887", "0.754", "0.817", "0.927", "0.847"])
    ws.append([])
    ws.append(["주의: 동일 100문항 A→B→F 순차 측정 → semantic_cache 영향 가능성."])
    ws.append(["새 평가셋 50문항 3-way 재검증으로 결론 확정."])


def build_report_md(args, a_ragas, b_ragas, f_ragas, a_judge, b_judge, f_judge, codex_text: str | None) -> str:
    a_total_ragas = mean([avg_metric(a_ragas, m) or 0 for m in RAGAS_METRICS])
    b_total_ragas = mean([avg_metric(b_ragas, m) or 0 for m in RAGAS_METRICS])
    f_total_ragas = mean([avg_metric(f_ragas, m) or 0 for m in RAGAS_METRICS])
    a_total_judge = avg_metric(a_judge, "total_score")
    b_total_judge = avg_metric(b_judge, "total_score")
    f_total_judge = avg_metric(f_judge, "total_score")
    a_f1 = avg_metric(a_judge, "kw_f1")
    b_f1 = avg_metric(b_judge, "kw_f1")
    f_f1 = avg_metric(f_judge, "kw_f1")

    lines = []
    lines.append("# Phase 2 — 새 평가셋 50문항 A vs B vs F 3-way 재검증")
    lines.append("")
    lines.append(f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    lines.append("")
    lines.append("## 종합 결론")
    lines.append("")
    lines.append(f"- RAGAS 4메트릭 평균 — A: **{a_total_ragas:.3f}**, B: **{b_total_ragas:.3f}**, F: **{f_total_ragas:.3f}** (최고: {winner_3way(a_total_ragas, b_total_ragas, f_total_ragas)})")
    if a_total_judge is not None:
        lines.append(f"- LLM-Judge 총점 (4~20) — A: **{a_total_judge:.2f}**, B: **{b_total_judge:.2f}**, F: **{f_total_judge:.2f}** (최고: {winner_3way(a_total_judge, b_total_judge, f_total_judge)})")
    if a_f1 is not None:
        lines.append(f"- 키워드 F1 — A: **{a_f1:.3f}**, B: **{b_f1:.3f}**, F: **{f_f1:.3f}** (최고: {winner_3way(a_f1, b_f1, f_f1)})")
    lines.append("")

    lines.append("## RAGAS 메트릭별 비교")
    lines.append("")
    lines.append("| 메트릭 | A | B | F | 최고 |")
    lines.append("|---|---:|---:|---:|---|")
    for m in RAGAS_METRICS:
        a_v = avg_metric(a_ragas, m)
        b_v = avg_metric(b_ragas, m)
        f_v = avg_metric(f_ragas, m)
        lines.append(f"| {m} | {a_v:.3f} | {b_v:.3f} | {f_v:.3f} | {winner_3way(a_v, b_v, f_v)} |")
    lines.append("")

    if a_judge:
        lines.append("## LLM-Judge 메트릭별 비교")
        lines.append("")
        lines.append("| 메트릭 | A | B | F | 최고 |")
        lines.append("|---|---:|---:|---:|---|")
        for m in JUDGE_METRICS:
            a_v = avg_metric(a_judge, m)
            b_v = avg_metric(b_judge, m)
            f_v = avg_metric(f_judge, m)
            lines.append(f"| {m} | {a_v:.2f} | {b_v:.2f} | {f_v:.2f} | {winner_3way(a_v, b_v, f_v)} |")
        lines.append("")

        lines.append("## L별 LLM-Judge 총점 분포")
        lines.append("")
        lines.append("| 난이도 | A | B | F | 최고 |")
        lines.append("|---|---:|---:|---:|---|")
        a_by_L = by_level(a_judge)
        b_by_L = by_level(b_judge)
        f_by_L = by_level(f_judge)
        for L in ("L1", "L2", "L3", "L4", "L5"):
            a_v = avg_metric(a_by_L.get(L, []), "total_score")
            b_v = avg_metric(b_by_L.get(L, []), "total_score")
            f_v = avg_metric(f_by_L.get(L, []), "total_score")
            lines.append(f"| {L} | {a_v:.2f} | {b_v:.2f} | {f_v:.2f} | {winner_3way(a_v, b_v, f_v)} |" if a_v is not None else f"| {L} | - | - | - | - |")
        lines.append("")

    lines.append("## 측정 조건")
    lines.append("")
    lines.append("- 평가셋: 참부모님 생애와 통일원리 문답 학습서 (50문항, ID 101~150)")
    lines.append("- L분포: L1~L5 각 10건씩 균등")
    lines.append("- 측정 순서: F → A → B (각 batch 사이 cache 비우기 + ensure)")
    lines.append("- 챗봇 토글: 'all' 봇 collection_main만 변경 (system_prompt/persona/search_tiers 동결)")
    lines.append("- 캐시: 측정 시작 시 `delete_collection` + `ensure_cache_collection`(빈 컬렉션)")
    lines.append("- 평가 모델: gemini-3.1-flash-lite-preview, temperature=0")
    lines.append("- Codex: OpenAI gpt-5-codex (consult mode, model_reasoning_effort=medium)")
    lines.append("")

    if codex_text:
        lines.append("## Codex 3-way 독립 검토 (10건 stratified)")
        lines.append("")
        lines.append(codex_text[:5000])
        if len(codex_text) > 5000:
            lines.append("")
            lines.append(f"... (전체 {len(codex_text)}자, 별도 md 참고)")
        lines.append("")

    lines.append("## 이전 100선 결과 (참고)")
    lines.append("")
    lines.append("| 옵션 | faithfulness | context_precision | context_recall | response_relevancy | 평균 |")
    lines.append("|---|---:|---:|---:|---:|---:|")
    lines.append("| A (v1) | 0.869 | 0.726 | 0.800 | 0.882 | 0.819 |")
    lines.append("| B (v2) | 0.864 | 0.746 | 0.784 | 0.890 | 0.821 |")
    lines.append("| F (v3) | 0.887 | 0.754 | 0.817 | 0.927 | **0.847** |")
    lines.append("")
    lines.append("> 동일 100문항 A→B→F 순차 측정으로 cache 영향 의심됨. 본 50문항 3-way 재검증으로 확인.")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ragas-a", required=True, type=Path)
    parser.add_argument("--ragas-b", required=True, type=Path)
    parser.add_argument("--ragas-f", required=True, type=Path)
    parser.add_argument("--judge-a-csv", required=True, type=Path)
    parser.add_argument("--judge-b-csv", required=True, type=Path)
    parser.add_argument("--judge-f-csv", required=True, type=Path)
    parser.add_argument("--codex-md", type=Path, default=None)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--report", required=True, type=Path)
    args = parser.parse_args()

    a_ragas = load_ragas(args.ragas_a)
    b_ragas = load_ragas(args.ragas_b)
    f_ragas = load_ragas(args.ragas_f)
    a_judge = load_judge_csv(args.judge_a_csv)
    b_judge = load_judge_csv(args.judge_b_csv)
    f_judge = load_judge_csv(args.judge_f_csv)
    codex_text = args.codex_md.read_text(encoding="utf-8") if (args.codex_md and args.codex_md.exists()) else None

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)
    write_ragas_sheet(wb.create_sheet("RAGAS 4메트릭"), a_ragas, b_ragas, f_ragas)
    write_judge_sheet(wb.create_sheet("LLM-Judge 정성"), a_judge, b_judge, f_judge)
    write_keyword_f1_sheet(wb.create_sheet("키워드 F1"), a_judge, b_judge, f_judge)
    if codex_text:
        write_codex_sheet(wb.create_sheet("Codex 검토"), codex_text)
    write_legacy_sheet(wb.create_sheet("이전 100선 (참고)"))

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)

    report = build_report_md(args, a_ragas, b_ragas, f_ragas, a_judge, b_judge, f_judge, codex_text)
    args.report.parent.mkdir(parents=True, exist_ok=True)
    args.report.write_text(report, encoding="utf-8")

    print(f"3-way 통합 xlsx: {args.output}")
    print(f"3-way 결론 보고서: {args.report}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
