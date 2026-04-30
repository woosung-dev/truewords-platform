"""
RAGAS 4메트릭 평가 본체.

Phase A의 sample_eval_pairs.py가 생성한 50건 시드(JSON)를 입력받아
RAGAS의 다음 4메트릭을 산출하고 xlsx로 저장한다.
- Faithfulness         : 답변이 contexts에 충실한가
- ContextPrecision     : contexts가 ground_truth와 얼마나 정렬되는가
- ContextRecall        : ground_truth를 contexts로 얼마나 회수하는가
- ResponseRelevancy    : 답변이 질문에 얼마나 관련 있는가

LLM (평가용): gemini-3.1-pro (임시 — docs/TODO.md "RAGAS 평가 LLM 환원" 참조).
인계 문서 §5 사전 결정은 Claude Haiku 4.5 였으나 Anthropic 크레딧 잔액 부족으로
일시적 Gemini 3.1 Pro 대체. 충전 후 환원 예정 (생성/평가 같은 모델 패밀리는
G-Eval LLM-self-bias 우려 — 잠정).
임베딩: gemini-embedding-001 (ResponseRelevancy의 cosine similarity용).

환경변수:
    GEMINI_API_KEY (필수): Gemini Pro 호출 + Gemini embedding 호출 모두 사용

사용:
    # 1) dry-run: 시드 로드 + RAGAS dataset 변환 검증만 (키 불필요)
    cd backend
    uv run --group eval python scripts/eval_ragas.py \\
        --seed ~/Downloads/ragas_eval_seed_50_<ts>.json \\
        --dry-run

    # 2) 실제 평가
    ANTHROPIC_API_KEY=... GEMINI_API_KEY=... \\
        uv run --group eval python scripts/eval_ragas.py \\
            --seed ~/Downloads/ragas_eval_seed_50_<ts>.json \\
            --output ~/Downloads/ragas_baseline_<ts>.xlsx

    # 3) 일부만 (디버깅)
    uv run --group eval python scripts/eval_ragas.py \\
        --seed <seed.json> --output <out.xlsx> --limit 5
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ragas는 평가 시점에만 import (dry-run에서는 무거운 의존성 회피)
DEFAULT_OUTPUT_DIR = Path.home() / "Downloads"

# 평가용 모델 (임시 — Anthropic 크레딧 충전 후 Claude Haiku 4.5 로 환원 예정).
# 환원 작업: docs/TODO.md "RAGAS 평가 LLM 환원" Blocked 항목 참조.
# gemini-3.1-pro-preview 는 RAGAS 평가에서 RPM throttling/응답 hang 다발로 사용 불가
# → gemini-2.5-pro (production 안정 모델) 로 fallback.
EVAL_LLM_MODEL = "gemini-2.5-pro"
EVAL_EMBEDDING_MODEL = "models/gemini-embedding-001"


@dataclass
class SeedItem:
    """sample_eval_pairs.py가 생성한 시드의 한 항목."""

    id: str
    source_file: str
    level: str
    category: str
    question: str
    answer: str
    ground_truth: str
    contexts: list[str]


def load_seed(path: Path) -> list[SeedItem]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        raise ValueError(f"시드 JSON은 list여야 함: {path}")
    items: list[SeedItem] = []
    for r in raw:
        items.append(
            SeedItem(
                id=r["id"],
                source_file=r["source_file"],
                level=r["level"],
                category=r["category"],
                question=r["question"],
                answer=r.get("answer", ""),
                ground_truth=r.get("ground_truth", ""),
                contexts=list(r.get("contexts", [])),
            )
        )
    return items


def to_ragas_samples(items: list[SeedItem]) -> Any:
    """SeedItem 리스트를 RAGAS EvaluationDataset으로 변환."""
    from ragas import EvaluationDataset, SingleTurnSample

    samples = [
        SingleTurnSample(
            user_input=it.question,
            retrieved_contexts=it.contexts,
            response=it.answer,
            reference=it.ground_truth,
        )
        for it in items
    ]
    return EvaluationDataset(samples=samples)


def build_evaluator() -> tuple[Any, Any]:
    """평가용 LLM/Embedding wrapper를 만든다. 환경변수 검증 포함."""
    if not os.getenv("GEMINI_API_KEY"):
        # backend/.env에서 로드 시도 — 운영 코드와 동일 키 사용
        env_path = Path(__file__).resolve().parent.parent / ".env"
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8").splitlines():
                if line.startswith("GEMINI_API_KEY="):
                    os.environ.setdefault("GEMINI_API_KEY", line.split("=", 1)[1].strip())
                    break
    if not os.getenv("GEMINI_API_KEY"):
        # Google API 키 환경변수 호환
        if os.getenv("GOOGLE_API_KEY"):
            os.environ["GEMINI_API_KEY"] = os.environ["GOOGLE_API_KEY"]
    if not os.getenv("GEMINI_API_KEY"):
        raise SystemExit(
            "GEMINI_API_KEY 미설정. backend/.env 또는 환경변수로 제공해야 합니다."
        )
    # langchain-google-genai는 GOOGLE_API_KEY를 본다
    os.environ.setdefault("GOOGLE_API_KEY", os.environ["GEMINI_API_KEY"])

    from langchain_google_genai import (
        ChatGoogleGenerativeAI,
        GoogleGenerativeAIEmbeddings,
    )
    from ragas.embeddings import LangchainEmbeddingsWrapper
    from ragas.llms import LangchainLLMWrapper

    # 첫 sanity (1m58s 5건 timeout 0) 와 동일 — 완전 default. 추가 파라미터(timeout, max_retries,
    # RunConfig)를 명시하면 hang 발생 (RAGAS 0.4.3 + langchain-google-genai 4.2.2 조합 이슈).
    eval_llm = LangchainLLMWrapper(
        ChatGoogleGenerativeAI(model=EVAL_LLM_MODEL, temperature=0)
    )
    eval_embeddings = LangchainEmbeddingsWrapper(
        GoogleGenerativeAIEmbeddings(model=EVAL_EMBEDDING_MODEL)
    )
    return eval_llm, eval_embeddings


def run_evaluation(
    items: list[SeedItem],
) -> dict[str, list[float | None]]:
    """RAGAS 평가 실행. 항목별 4메트릭 점수 dict 반환."""
    import time as _t
    def _log(msg: str) -> None:
        print(f"[ragas-debug {_t.strftime('%H:%M:%S')}] {msg}", flush=True)

    _log("step 1: import ragas")
    from ragas import evaluate
    from ragas.metrics import (
        ContextPrecision,
        ContextRecall,
        Faithfulness,
        ResponseRelevancy,
    )
    _log("step 2: build_evaluator()")
    eval_llm, eval_embeddings = build_evaluator()
    _log("step 3: to_ragas_samples()")
    dataset = to_ragas_samples(items)
    _log(f"step 4: build metrics (items={len(items)})")
    # strictness=1: Gemini Pro 가 multiple candidates 미지원 → multi-sampling 메트릭은
    # 단일 candidate 로 계산. Claude Haiku 환원 시 strictness 기본값(3) 으로 복원 가능.
    metrics = [
        Faithfulness(),
        ContextPrecision(),
        ContextRecall(),
        ResponseRelevancy(strictness=1),
    ]
    _log("step 5: evaluate() 호출 시작 — allow_nest_asyncio=True 추가")
    # ragas 0.4.3 hang 회피 — asyncio nested loop 허용
    result = evaluate(
        dataset=dataset,
        metrics=metrics,
        llm=eval_llm,
        embeddings=eval_embeddings,
        show_progress=True,
        allow_nest_asyncio=True,
        batch_size=1,
    )
    _log("step 6: evaluate() 완료")
    # result.scores는 list[dict[metric_name, float|nan]]
    scores: dict[str, list[float | None]] = {
        "faithfulness": [],
        "context_precision": [],
        "context_recall": [],
        "answer_relevancy": [],
    }
    for row in result.scores:
        for metric_name in scores.keys():
            v = row.get(metric_name)
            try:
                fv = float(v) if v is not None else None
                if fv is not None and (fv != fv):  # NaN
                    fv = None
                scores[metric_name].append(fv)
            except (TypeError, ValueError):
                scores[metric_name].append(None)
    return scores


def write_xlsx(
    items: list[SeedItem],
    scores: dict[str, list[float | None]],
    out_path: Path,
) -> None:
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "ragas"

    headers = [
        "id",
        "source_file",
        "level",
        "category",
        "question",
        "ground_truth",
        "our_answer",
        "n_contexts",
        "faithfulness",
        "context_precision",
        "context_recall",
        "answer_relevancy",
        "avg",
    ]
    ws.append(headers)
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[1]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, it in enumerate(items):
        f = scores["faithfulness"][i]
        cp = scores["context_precision"][i]
        cr = scores["context_recall"][i]
        ar = scores["answer_relevancy"][i]
        valid = [v for v in (f, cp, cr, ar) if v is not None]
        avg = sum(valid) / len(valid) if valid else None
        ws.append(
            [
                it.id,
                it.source_file,
                it.level,
                it.category,
                it.question,
                it.ground_truth,
                it.answer,
                len(it.contexts),
                f,
                cp,
                cr,
                ar,
                avg,
            ]
        )

    widths = [28, 32, 22, 18, 40, 50, 50, 8, 12, 14, 12, 14, 8]
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        # 점수 셀 소수점 3자리
        for col_idx in (9, 10, 11, 12, 13):
            row[col_idx - 1].number_format = "0.000"
    ws.freeze_panes = "A2"

    # 요약 시트
    ws_sum = wb.create_sheet("summary")
    ws_sum.append(["metric", "n", "mean", "min", "max"])
    for cell in ws_sum[1]:
        cell.font = bold
        cell.fill = fill
    for metric_name in ("faithfulness", "context_precision", "context_recall", "answer_relevancy"):
        valid = [v for v in scores[metric_name] if v is not None]
        if valid:
            ws_sum.append([metric_name, len(valid), sum(valid) / len(valid), min(valid), max(valid)])
        else:
            ws_sum.append([metric_name, 0, None, None, None])
    for col_idx in range(1, 6):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = 22

    wb.save(out_path)


def print_summary(scores: dict[str, list[float | None]]) -> None:
    print("\n=== RAGAS Summary ===")
    for metric_name, values in scores.items():
        valid = [v for v in values if v is not None]
        if valid:
            mean = sum(valid) / len(valid)
            print(f"  {metric_name:24s} mean={mean:.3f}  n={len(valid)}/{len(values)}")
        else:
            print(f"  {metric_name:24s} (no valid scores)")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--seed",
        type=Path,
        required=True,
        help="sample_eval_pairs.py가 생성한 JSON 시드 경로",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help=f"출력 xlsx 경로 (default: {DEFAULT_OUTPUT_DIR}/ragas_<timestamp>.xlsx)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="앞에서부터 N건만 평가 (디버깅용)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="실제 RAGAS 호출 없이 시드 로드 + dataset 변환만 검증",
    )
    args = parser.parse_args()

    items = load_seed(args.seed)
    if args.limit:
        items = items[: args.limit]
    print(f"[load] {args.seed} → {len(items)}건")
    print(f"  contexts 평균: {sum(len(i.contexts) for i in items) / max(len(items), 1):.2f}")

    if args.dry_run:
        # ragas 임포트만 가능한지 + dataset 변환 가능한지 검증
        ds = to_ragas_samples(items)
        print(f"[dry-run] EvaluationDataset 생성 OK: {len(ds.samples)} samples")
        first = ds.samples[0]
        print(f"  첫 샘플 user_input(앞 50자): {first.user_input[:50]!r}...")
        print(f"  첫 샘플 retrieved_contexts: {len(first.retrieved_contexts or [])}건")
        print(f"  첫 샘플 reference(앞 50자): {(first.reference or '')[:50]!r}...")
        print("[dry-run] 모든 검증 통과. 실제 평가는 --dry-run 없이 재실행하세요.")
        return

    scores = run_evaluation(items)
    print_summary(scores)

    if args.output is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        args.output = DEFAULT_OUTPUT_DIR / f"ragas_{ts}.xlsx"
    write_xlsx(items, scores, args.output)
    print(f"\n→ {args.output}")


if __name__ == "__main__":
    sys.exit(main())
