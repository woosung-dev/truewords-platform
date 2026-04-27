"""3 개의 RAGAS xlsx (baseline / +action2 / +action1+2) 를 횡렬로 머지해 비교 xlsx 생성.

eval_ragas.py 가 출력한 xlsx 3개를 받아 id 기준 join, 4메트릭 × 3시나리오의
값을 한 시트에 펼쳐서 보고서 작성에 사용한다.

각 시나리오 컬럼 그룹:
    {prefix}_F   = faithfulness
    {prefix}_CP  = context_precision
    {prefix}_CR  = context_recall
    {prefix}_AR  = answer_relevancy
    {prefix}_avg = 4메트릭 평균

추가 컬럼:
    delta12_avg = (action1+2 avg) - (baseline avg)
    delta12_F   = (action1+2 F) - (baseline F)   ... 4메트릭 + avg
    delta2_avg  = (action2 avg) - (baseline avg)
    delta1_avg  = (action1+2 avg) - (action2 avg)  # 액션 1 단독 효과

사용:
    cd backend
    uv run --group eval python scripts/merge_ragas_3way.py \\
        --baseline ~/Downloads/ragas_baseline_<ts>.xlsx \\
        --action2 ~/Downloads/ragas_action2_<ts>.xlsx \\
        --action1plus2 ~/Downloads/ragas_action1plus2_<ts>.xlsx \\
        --output ~/Downloads/ragas_3way_<ts>.xlsx
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

METRICS = ("faithfulness", "context_precision", "context_recall", "answer_relevancy")
METRIC_SHORT = {"faithfulness": "F", "context_precision": "CP", "context_recall": "CR", "answer_relevancy": "AR"}


def load_ragas_xlsx(path: Path) -> dict[str, dict[str, float | None]]:
    """eval_ragas.py 출력 xlsx 의 한 행을 id → {metric: value} dict 로 변환."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"활성 시트 없음: {path}")
    rows = ws.iter_rows(values_only=True)
    header = [c for c in next(rows)]
    # 헤더 위치
    try:
        col_id = header.index("id")
        col_idx = {m: header.index(m) for m in METRICS}
    except ValueError as e:
        raise ValueError(f"{path} 컬럼 누락: {e}\n  헤더: {header}")
    out: dict[str, dict[str, float | None]] = {}
    for r in rows:
        if not r or not r[col_id]:
            continue
        rid = str(r[col_id])
        scores: dict[str, float | None] = {}
        for m in METRICS:
            v = r[col_idx[m]]
            try:
                fv = float(v) if v is not None else None
                if fv is not None and fv != fv:  # NaN
                    fv = None
                scores[m] = fv
            except (TypeError, ValueError):
                scores[m] = None
        out[rid] = scores
    wb.close()
    return out


def load_metadata(path: Path) -> dict[str, dict[str, Any]]:
    """RAGAS xlsx 의 metadata 컬럼(id/level/category/question/ground_truth/our_answer) 만 추출."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"활성 시트 없음: {path}")
    rows = ws.iter_rows(values_only=True)
    header = [c for c in next(rows)]
    cols = {}
    for name in ("id", "level", "category", "question", "ground_truth", "our_answer", "source_file"):
        if name in header:
            cols[name] = header.index(name)
    out: dict[str, dict[str, Any]] = {}
    for r in rows:
        if not r or "id" not in cols or not r[cols["id"]]:
            continue
        rid = str(r[cols["id"]])
        out[rid] = {k: r[idx] for k, idx in cols.items()}
    wb.close()
    return out


def avg_of(scores: dict[str, float | None]) -> float | None:
    valid = [v for v in scores.values() if v is not None]
    return sum(valid) / len(valid) if valid else None


def safe_diff(a: float | None, b: float | None) -> float | None:
    if a is None or b is None:
        return None
    return a - b


def build_xlsx(
    baseline: dict[str, dict[str, float | None]],
    action2: dict[str, dict[str, float | None]],
    a1plus2: dict[str, dict[str, float | None]],
    metadata: dict[str, dict[str, Any]],
    out_path: Path,
) -> None:
    wb = openpyxl.Workbook()
    ws: Worksheet = wb.active  # type: ignore[assignment]
    ws.title = "3way"

    headers = ["id", "source_file", "level", "category", "question", "ground_truth"]
    for prefix in ("base", "act2", "act12"):
        for m in METRICS:
            headers.append(f"{prefix}_{METRIC_SHORT[m]}")
        headers.append(f"{prefix}_avg")
    headers.extend(
        [
            "delta_act2_F", "delta_act2_CP", "delta_act2_CR", "delta_act2_AR", "delta_act2_avg",
            "delta_act12_F", "delta_act12_CP", "delta_act12_CR", "delta_act12_AR", "delta_act12_avg",
            "delta_act1_only_avg",
        ]
    )
    ws.append(headers)
    bold = Font(bold=True)
    fill_meta = PatternFill("solid", fgColor="DDDDDD")
    fill_base = PatternFill("solid", fgColor="FFE4B5")  # 살구
    fill_act2 = PatternFill("solid", fgColor="FFFACD")  # 연노랑
    fill_act12 = PatternFill("solid", fgColor="C7E9C0")  # 연초록
    fill_delta = PatternFill("solid", fgColor="DCE7F5")  # 연파랑
    for cell in ws[1]:
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # 색상 (인덱스 기반)
    # 6개 metadata + base 5 + act2 5 + act12 5 + delta 11 = 32
    for col_idx in range(1, 7):
        ws.cell(row=1, column=col_idx).fill = fill_meta
    for col_idx in range(7, 12):
        ws.cell(row=1, column=col_idx).fill = fill_base
    for col_idx in range(12, 17):
        ws.cell(row=1, column=col_idx).fill = fill_act2
    for col_idx in range(17, 22):
        ws.cell(row=1, column=col_idx).fill = fill_act12
    for col_idx in range(22, 33):
        ws.cell(row=1, column=col_idx).fill = fill_delta

    all_ids = sorted(set(baseline) | set(action2) | set(a1plus2))
    for rid in all_ids:
        meta = metadata.get(rid, {})
        b = baseline.get(rid, {m: None for m in METRICS})
        a2 = action2.get(rid, {m: None for m in METRICS})
        a12 = a1plus2.get(rid, {m: None for m in METRICS})
        b_avg = avg_of(b)
        a2_avg = avg_of(a2)
        a12_avg = avg_of(a12)
        row = [
            rid,
            meta.get("source_file"),
            meta.get("level"),
            meta.get("category"),
            meta.get("question"),
            meta.get("ground_truth"),
            *(b[m] for m in METRICS), b_avg,
            *(a2[m] for m in METRICS), a2_avg,
            *(a12[m] for m in METRICS), a12_avg,
            *(safe_diff(a2[m], b[m]) for m in METRICS), safe_diff(a2_avg, b_avg),
            *(safe_diff(a12[m], b[m]) for m in METRICS), safe_diff(a12_avg, b_avg),
            safe_diff(a12_avg, a2_avg),
        ]
        ws.append(row)

    # 컬럼 너비 + 점수 셀 포맷
    widths = [28, 32, 22, 18, 40, 50] + [10] * 26
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row_idx in range(2, ws.max_row + 1):
        for col_idx in range(7, 33):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = "0.000;[Red]-0.000"
            cell.alignment = Alignment(wrap_text=False, vertical="top")
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "G2"

    # 요약 시트
    ws_sum = wb.create_sheet("summary")
    ws_sum.append(["scenario", "n", "F_mean", "CP_mean", "CR_mean", "AR_mean", "avg_mean"])
    for cell in ws_sum[1]:
        cell.font = bold
        cell.fill = fill_meta
    for label, scores_map in [
        ("baseline", baseline),
        ("action2", action2),
        ("action1+2", a1plus2),
    ]:
        n = len(scores_map)
        per_metric_means: list[float | None] = []
        for m in METRICS:
            vals = [s[m] for s in scores_map.values() if s.get(m) is not None]
            per_metric_means.append(sum(vals) / len(vals) if vals else None)
        avg_vals = [avg_of(s) for s in scores_map.values()]
        avg_valid = [v for v in avg_vals if v is not None]
        avg_mean = sum(avg_valid) / len(avg_valid) if avg_valid else None
        ws_sum.append([label, n, *per_metric_means, avg_mean])
    for col_idx in range(1, 8):
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = 16
    for row_idx in range(2, ws_sum.max_row + 1):
        for col_idx in range(3, 8):
            ws_sum.cell(row=row_idx, column=col_idx).number_format = "0.000"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--baseline", type=Path, required=True, help="ragas_baseline_*.xlsx")
    p.add_argument("--action2", type=Path, required=True, help="ragas_action2_*.xlsx")
    p.add_argument("--action1plus2", type=Path, required=True, help="ragas_action1plus2_*.xlsx")
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        help="기본: ~/Downloads/ragas_3way_<ts>.xlsx",
    )
    args = p.parse_args()

    print(f"[load] baseline    {args.baseline.name}")
    baseline = load_ragas_xlsx(args.baseline)
    print(f"[load] action2     {args.action2.name}")
    action2 = load_ragas_xlsx(args.action2)
    print(f"[load] action1+2   {args.action1plus2.name}")
    a1plus2 = load_ragas_xlsx(args.action1plus2)
    metadata = load_metadata(args.baseline)
    metadata.update(load_metadata(args.action2))
    metadata.update(load_metadata(args.action1plus2))

    if args.output is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        args.output = Path.home() / "Downloads" / f"ragas_3way_{ts}.xlsx"

    build_xlsx(baseline, action2, a1plus2, metadata, args.output)
    print(f"\n→ {args.output}")
    print(f"  baseline  ids: {len(baseline)}")
    print(f"  action2   ids: {len(action2)}")
    print(f"  action1+2 ids: {len(a1plus2)}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
