"""베이스라인 + 튜닝 평가 결과 두 xlsx를 한 시트에 나란히 머지.

사용 예:
    PYTHONPATH=. uv run python scripts/merge_eval_compare.py \
        --baseline /Users/woosung/Downloads/천일국섭리_베이스라인.xlsx \
        --tuned    /Users/woosung/Downloads/천일국섭리_튜닝후.xlsx \
        --output   /Users/woosung/Downloads/천일국섭리_비교.xlsx

생성 컬럼 (좌→우):
    번호 | 2번째컬럼 | 카테고리 | 질문 | 모범답변 | 참고키워드 |
    [BASE] 답변 | [BASE] 참고1 | [BASE] 참고2 | [BASE] 참고3 |
    [TUNE] 답변 | [TUNE] 참고1 | [TUNE] 참고2 | [TUNE] 참고3
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def fix_num(v):
    return int(v) if isinstance(v, str) and v.isdigit() else v


def load_indexed(path: Path) -> tuple[list[str], dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c) if c is not None else "" for c in rows[0]]
    indexed: dict = {}
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        n = fix_num(r[0])
        indexed[n] = r
    return header, indexed


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--baseline", required=True, type=Path)
    p.add_argument("--tuned", required=True, type=Path)
    p.add_argument("--output", required=True, type=Path)
    args = p.parse_args()

    base_header, base = load_indexed(args.baseline)
    tune_header, tune = load_indexed(args.tuned)

    # 공통 번호 정렬
    nums = sorted(set(base.keys()) | set(tune.keys()), key=lambda x: (isinstance(x, str), x))

    wb = Workbook()
    ws = wb.active
    ws.title = "베이스vs튜닝"

    # 헤더 (앞 6컬럼은 baseline 기준)
    new_header = base_header[:6] + [
        "[베이스] 답변", "[베이스] 참고1", "[베이스] 참고2", "[베이스] 참고3",
        "[튜닝] 답변",   "[튜닝] 참고1",   "[튜닝] 참고2",   "[튜닝] 참고3",
    ]
    ws.append(new_header)

    bold = Font(bold=True)
    base_fill = PatternFill("solid", fgColor="FFF7ED")  # 연한 주황
    tune_fill = PatternFill("solid", fgColor="ECFDF5")  # 연한 초록
    for i, cell in enumerate(ws[1], 1):
        cell.font = bold
        if i in (7, 8, 9, 10):
            cell.fill = base_fill
        elif i in (11, 12, 13, 14):
            cell.fill = tune_fill

    for n in nums:
        b = base.get(n)
        t = tune.get(n)
        # 앞 6컬럼은 baseline 우선 (없으면 tuned)
        front = list((b or t)[:6])
        front[0] = n  # 번호 정수화
        b_ans = (b[6] if b and len(b) > 6 else "") or ""
        b_s1 = (b[7] if b and len(b) > 7 else "") or ""
        b_s2 = (b[8] if b and len(b) > 8 else "") or ""
        b_s3 = (b[9] if b and len(b) > 9 else "") or ""
        t_ans = (t[6] if t and len(t) > 6 else "") or ""
        t_s1 = (t[7] if t and len(t) > 7 else "") or ""
        t_s2 = (t[8] if t and len(t) > 8 else "") or ""
        t_s3 = (t[9] if t and len(t) > 9 else "") or ""
        ws.append(front + [b_ans, b_s1, b_s2, b_s3, t_ans, t_s1, t_s2, t_s3])

    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
        for i in (7, 8, 9, 10):
            row[i - 1].fill = base_fill
        for i in (11, 12, 13, 14):
            row[i - 1].fill = tune_fill

    widths = [6, 14, 14, 50, 50, 22, 60, 50, 50, 50, 60, 50, 50, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print(f"머지 완료: {args.output} ({len(nums)}건)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
