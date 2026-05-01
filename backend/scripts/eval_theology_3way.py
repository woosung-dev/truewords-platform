"""Phase 4 — 5권 신학/원리 3-way 측정 (Recursive vs Hierarchical vs Contextual).

기존 eval_notebooklm_qa.py 가 chatbot_id 기반 /chat 호출인데, 본 PoC 는
3개 컬렉션 직접 비교가 필요해서 chat 파이프라인 일부를 직접 호출 (검색 +
rerank + 생성, safety/cache 우회).

사용:
    PYTHONPATH=. uv run python scripts/eval_theology_3way.py \\
        --input ~/Downloads/통일원리_평화사상_normalized.xlsx \\
        --collection theology_poc_recursive \\
        --output ~/Downloads/notebooklm_qa_theology_recursive_n100_<TS>.xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import sys
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from src.chat.generator import generate_answer
from src.chat.prompt import DEFAULT_SYSTEM_PROMPT
from src.chatbot.runtime_config import GenerationConfig
from src.common.gemini import embed_dense_query
from src.pipeline.embedder import embed_sparse_async
from src.qdrant_client import get_raw_client
from src.search.hybrid import SearchResult, hybrid_search
from src.search.reranker import rerank


async def run_one(
    question: str,
    collection: str,
    top_k_search: int = 50,
    top_k_rerank: int = 10,
) -> tuple[str, list[SearchResult]]:
    """단일 질문 → 검색 → rerank → 생성 → (answer, top10 results)."""
    client = get_raw_client()
    dense = await embed_dense_query(question)
    sparse = await embed_sparse_async(question)
    results = await hybrid_search(
        client,
        question,
        top_k=top_k_search,
        dense_embedding=dense,
        sparse_embedding=sparse,
        collection_name=collection,
    )
    if not results:
        return "[오류] 검색 결과 없음", []
    reranked = await rerank(question, results, top_k=top_k_rerank)

    # Hierarchical 의 경우, build_context_prompt 가 result.text 를 사용하므로
    # parent_text 가 있으면 SearchResult.text 를 parent 로 교체 (in-place)
    for r in reranked:
        if r.parent_text:
            r.text = r.parent_text  # type: ignore[misc]

    gen_config = GenerationConfig(
        system_prompt=DEFAULT_SYSTEM_PROMPT,
        persona_name="신학/원리 전문 봇",
    )
    answer = await generate_answer(
        query=question,
        results=reranked,
        generation_config=gen_config,
    )
    return answer, reranked


def write_output(rows: list[dict], out_path: Path, ans_col: str) -> None:
    headers = [
        "번호", "난이도(Level)", "카테고리", "테스트용 질문",
        "봇 모범 답변", "참고 키워드",
        ans_col, "참고1", "참고2", "참고3", "세션ID",
    ]
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = bold
    for r_idx, r in enumerate(rows, 2):
        for c_idx, h in enumerate(headers, 1):
            v = r.get(h, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=v)
            cell.alignment = wrap
    widths = [6, 14, 14, 40, 40, 18, 60, 30, 30, 30, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


async def amain(args: argparse.Namespace) -> int:
    wb = load_workbook(args.input, data_only=True)
    ws = wb.active
    rows_raw = list(ws.iter_rows(values_only=True))
    if not rows_raw:
        raise SystemExit("입력 xlsx 비어있음")

    ans_col = f"우리 답변({args.label})"
    out_rows: list[dict] = []
    failures = 0
    started = time.time()

    sleep_s = 1.0 / max(args.rate_per_sec, 0.01)

    for i, raw in enumerate(rows_raw[1:], 1):
        if args.limit and i > args.limit:
            break
        if not raw or all(c is None or c == "" for c in raw):
            continue
        num, level, category, question, gt, kw = raw[:6]
        question = (question or "").strip()
        if not question:
            continue
        preview = question.replace("\n", " ")[:60]
        print(f"[{i}/100] #{num} {(level or '')[:8]} | {preview}", flush=True)
        try:
            answer, results = await run_one(
                question, collection=args.collection,
                top_k_search=args.top_k_search,
                top_k_rerank=args.top_k_rerank,
            )
        except Exception as e:
            answer = f"[오류] {type(e).__name__}: {e}"
            results = []
            failures += 1

        # 참고1/2/3 — Hierarchical 의 경우 parent_text 우선
        contexts = []
        for r in results[:3]:
            text = r.parent_text or r.text
            contexts.append(f"[{r.volume}] (score={r.score:.3f}, source={r.source})\n{text[:300]}")
        while len(contexts) < 3:
            contexts.append("")

        ans_preview = (answer or "").replace("\n", " ")[:80]
        print(f"   → results={len(results)} | {ans_preview}", flush=True)

        out_rows.append({
            "번호": num,
            "난이도(Level)": level,
            "카테고리": category,
            "테스트용 질문": question,
            "봇 모범 답변": gt,
            "참고 키워드": kw,
            ans_col: answer,
            "참고1": contexts[0],
            "참고2": contexts[1],
            "참고3": contexts[2],
            "세션ID": "",  # PoC — chat session 없음
        })

        if i % 25 == 0 or (args.limit and i == args.limit):
            cp = args.output.with_name(args.output.stem + ".checkpoint.xlsx")
            write_output(out_rows, cp, ans_col=ans_col)
            print(f"   ✓ checkpoint 저장 ({i}건)", flush=True)

        await asyncio.sleep(sleep_s)

    write_output(out_rows, args.output, ans_col=ans_col)
    elapsed = time.time() - started
    print()
    print(f"=== 완료 ({args.collection}) ===")
    print(f"성공: {len(out_rows) - failures} / 실패: {failures}")
    print(f"소요: {elapsed:.1f}s")
    print(f"결과: {args.output}")
    return 0


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", required=True, type=Path)
    p.add_argument("--collection", required=True)
    p.add_argument("--label", default=None,
                   help="답변 컬럼명 suffix (기본 collection 마지막 토큰)")
    p.add_argument("--output", required=True, type=Path)
    p.add_argument("--rate-per-sec", type=float, default=0.5)
    p.add_argument("--limit", type=int, default=None)
    p.add_argument("--top-k-search", type=int, default=50)
    p.add_argument("--top-k-rerank", type=int, default=10)
    args = p.parse_args()
    if args.label is None:
        args.label = args.collection.replace("theology_poc_", "")
    return asyncio.run(amain(args))


if __name__ == "__main__":
    sys.exit(main())
