"""NotebookLM Q&A 두 Excel 파일을 합쳐 단일 챗봇 버전으로 일괄 평가.

사용 예:
    cd backend
    PYTHONPATH=. uv run python scripts/eval_notebooklm_qa.py \
      --light  "/Users/woosung/Downloads/Light RAG 성능 평가를 위한 단계별 테스트 Q&A 50선.xlsx" \
      --level5 "/Users/woosung/Downloads/통일원리 및 RAG 성능 평가용 5단계 숙련도 테스트 세트.xlsx" \
      --output "/Users/woosung/Downloads/notebooklm_qa_전체검색봇_평가.xlsx" \
      --chatbot-id all \
      --api-base http://localhost:8000

옵션:
    --limit N     상위 N건만 (스모크 테스트용)
    --rate-per-sec  분당 요청 한도 방어 (default 0.3 = 3.3s 간격)

동작:
    1. 두 xlsx 헤더 동일 가정: [번호 | 난이도(Level) | 카테고리 | 테스트용 질문 | 봇 모범 답변 | 참고 키워드]
    2. POST {api_base}/chat {"query": <질문>, "chatbot_id": <id>} 직렬 호출
    3. 응답에서 answer + sources[:3] 추출
    4. 결과 xlsx에 5개 신규 컬럼 추가:
       - 우리 답변(<chatbot_id>)
       - 참고1, 참고2, 참고3
       - 세션ID
    5. 50건마다 임시 .xlsx 체크포인트 저장
"""
from __future__ import annotations

import argparse
import sys
import time
from dataclasses import dataclass
from pathlib import Path

import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

EXPECTED_HEADER = [
    "번호", "난이도(Level)", "카테고리", "테스트용 질문", "봇 모범 답변", "참고 키워드",
]


@dataclass
class QARow:
    번호: int | str
    난이도: str
    카테고리: str
    질문: str
    모범답변: str
    참고키워드: str


def load_xlsx_rows(path: Path) -> list[QARow]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"{path} 비어 있음")
    header = [(c or "").strip() for c in rows[0]]
    if header[: len(EXPECTED_HEADER)] != EXPECTED_HEADER:
        raise SystemExit(
            f"{path} 헤더 불일치\n  기대: {EXPECTED_HEADER}\n  실제: {header}"
        )
    out: list[QARow] = []
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        out.append(
            QARow(
                번호=r[0] if r[0] is not None else "",
                난이도=str(r[1] or ""),
                카테고리=str(r[2] or ""),
                질문=str(r[3] or ""),
                모범답변=str(r[4] or ""),
                참고키워드=str(r[5] or ""),
            )
        )
    return out


def format_source_cell(src: dict) -> str:
    vol = src.get("volume", "") or ""
    score = src.get("score", 0) or 0
    source_code = src.get("source", "") or ""
    text = (src.get("text", "") or "").replace("\r", " ").strip()
    return f"[{vol}] (score={score:.3f}, source={source_code})\n{text}"


def call_chat(
    client: httpx.Client, *, api_base: str, query: str, chatbot_id: str
) -> tuple[str, list[dict], str]:
    url = f"{api_base.rstrip('/')}/chat"
    try:
        r = client.post(
            url,
            json={"query": query, "chatbot_id": chatbot_id},
            timeout=180.0,
        )
    except Exception as e:
        return f"[오류] 네트워크: {e}", [], ""
    if r.status_code != 200:
        return f"[오류] HTTP {r.status_code}: {r.text[:300]}", [], ""
    d = r.json()
    return (
        d.get("answer", "") or "",
        list(d.get("sources", []) or []),
        str(d.get("session_id", "") or ""),
    )


def write_output(
    out_path: Path,
    qrows: list[QARow],
    answers: list[str],
    sources_list: list[list[dict]],
    session_ids: list[str],
    chatbot_id: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "전체평가"

    new_header = list(EXPECTED_HEADER) + [
        f"우리 답변({chatbot_id})",
        "참고1",
        "참고2",
        "참고3",
        "세션ID",
    ]
    ws.append(new_header)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold

    wrap = Alignment(wrap_text=True, vertical="top")
    for q, ans, srcs, sid in zip(qrows, answers, sources_list, session_ids):
        s1 = format_source_cell(srcs[0]) if len(srcs) >= 1 else ""
        s2 = format_source_cell(srcs[1]) if len(srcs) >= 2 else ""
        s3 = format_source_cell(srcs[2]) if len(srcs) >= 3 else ""
        ws.append([
            q.번호, q.난이도, q.카테고리, q.질문, q.모범답변, q.참고키워드,
            ans, s1, s2, s3, sid,
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap

    widths = [6, 18, 14, 50, 50, 22, 60, 50, 50, 50, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--light", required=True, type=Path)
    p.add_argument("--level5", required=True, type=Path)
    p.add_argument("--output", required=True, type=Path)
    p.add_argument("--chatbot-id", default="all")
    p.add_argument("--api-base", default="http://localhost:8000")
    p.add_argument("--rate-per-sec", type=float, default=0.3)
    p.add_argument("--limit", type=int, default=None)
    args = p.parse_args()

    qrows = load_xlsx_rows(args.light) + load_xlsx_rows(args.level5)
    if args.limit:
        qrows = qrows[: args.limit]
    print(f"총 {len(qrows)}건 호출 예정 (chatbot_id={args.chatbot_id}, api={args.api_base})")

    sleep_s = 1.0 / max(args.rate_per_sec, 0.01)
    answers: list[str] = []
    sources_list: list[list[dict]] = []
    session_ids: list[str] = []

    checkpoint = args.output.with_name(args.output.stem + ".checkpoint.xlsx")

    failures = 0
    started = time.time()
    with httpx.Client() as client:
        for i, q in enumerate(qrows, 1):
            preview = q.질문.replace("\n", " ")[:60]
            print(f"[{i}/{len(qrows)}] #{q.번호} {q.난이도[:8]} | {preview}")
            ans, srcs, sid = call_chat(
                client,
                api_base=args.api_base,
                query=q.질문,
                chatbot_id=args.chatbot_id,
            )
            answers.append(ans)
            sources_list.append(srcs)
            session_ids.append(sid)
            ans_preview = (ans or "").replace("\n", " ")[:80]
            print(f"   → sources={len(srcs)} | {ans_preview}")
            if ans.startswith("[오류]"):
                failures += 1

            # 체크포인트 (50건마다 또는 마지막)
            if i % 50 == 0 or i == len(qrows):
                pad_q = qrows[:i]
                write_output(
                    checkpoint, pad_q, answers, sources_list, session_ids, args.chatbot_id
                )
                print(f"   ✓ checkpoint 저장: {checkpoint}")

            if i < len(qrows):
                time.sleep(sleep_s)

    write_output(
        args.output, qrows, answers, sources_list, session_ids, args.chatbot_id
    )
    if checkpoint.exists() and checkpoint != args.output:
        checkpoint.unlink()

    elapsed = time.time() - started
    print(
        f"\n완료: {len(qrows)}건 / 실패 {failures} / 소요 {elapsed:.1f}s\n"
        f"결과: {args.output}"
    )
    return 0 if failures == 0 else 2


if __name__ == "__main__":
    sys.exit(main())
