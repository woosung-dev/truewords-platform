"""Phase 3 메타데이터 필터 PoC Codex 정성 비교 markdown 생성.

v5 baseline (R2) vs v5 + 메타데이터 필터 비교 — 두 영역으로 구성:
  Section 1: volume_num 필터 적용 9건 (직접 영향 케이스)
  Section 2: L1~L5 각 1건 control (필터 미적용, 부작용 검증)

사용:
    PYTHONPATH=. uv run python scripts/build_phase3_codex_compare.py \\
      --xlsx-baseline ~/Downloads/notebooklm_qa_v5_n100_LR_round2_20260430_162131.xlsx \\
      --xlsx-phase3 ../tmp_match/phase3/eval_v5_meta_light100_20260430_2327.xlsx \\
      --output ../tmp_match/phase3/codex_compare_phase3.md
"""
from __future__ import annotations

import argparse
import random
from pathlib import Path

from openpyxl import load_workbook

from src.search.metadata_extractor import extract_query_metadata


def detect_level(level_raw: str) -> str:
    for L in ("L1", "L2", "L3", "L4", "L5"):
        if L in level_raw:
            return L
    return "기타"


def find_answer_col(headers: list[str]) -> str | None:
    for h in headers:
        if "우리 답변" in h:
            return h
    return None


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        out.append({headers[i]: r[i] if i < len(r) else None for i in range(len(headers))})
    return out


def extract_pair(row: dict, ans_col: str) -> tuple[str, list[str]]:
    answer = str(row.get(ans_col, "") or "")
    contexts = []
    for col in ("참고1", "참고2", "참고3"):
        c = row.get(col)
        if c:
            contexts.append(str(c)[:280])
    return answer, contexts


def render_case(idx: int, q_row: dict, b_ans: str, b_ctx: list[str], p_ans: str, p_ctx: list[str], section: str) -> list[str]:
    L = detect_level(str(q_row.get("난이도(Level)", "")))
    cat = str(q_row.get("카테고리", "") or "")
    q = (q_row.get("테스트용 질문") or "").strip()
    gt = str(q_row.get("봇 모범 답변", "") or "")
    kw = str(q_row.get("참고 키워드", "") or "")
    meta = extract_query_metadata(q)
    meta_str = f"`{meta}`" if meta else "—"

    out = [
        f"## [{section}] 사례 {idx} — {L} ({cat})",
        "",
        f"**질문**: {q}",
        "",
        f"**모범답변**: {gt}",
        "",
        f"**참고 키워드**: {kw}",
        "",
        f"**추출된 메타데이터**: {meta_str}",
        "",
        "### v5 baseline 답변 (필터 미적용)",
        "",
        b_ans,
        "",
    ]
    if b_ctx:
        out.append("**참고**:")
        for i, c in enumerate(b_ctx, 1):
            out.append(f"{i}. {c}")
        out.append("")

    out.extend([
        "### Phase 3 답변 (메타데이터 필터 적용)",
        "",
        p_ans,
        "",
    ])
    if p_ctx:
        out.append("**참고**:")
        for i, c in enumerate(p_ctx, 1):
            out.append(f"{i}. {c}")
        out.append("")
    out.append("---")
    out.append("")
    return out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-baseline", required=True, type=Path, help="v5 baseline xlsx (예: round2)")
    parser.add_argument("--xlsx-phase3", required=True, type=Path, help="Phase 3 측정 xlsx")
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--seed", type=int, default=20260430)
    args = parser.parse_args()

    base_rows = load_rows(args.xlsx_baseline)
    p3_rows = load_rows(args.xlsx_phase3)
    base_ans = find_answer_col(list(base_rows[0].keys()))
    p3_ans = find_answer_col(list(p3_rows[0].keys()))

    base_by_q = {(r.get("테스트용 질문") or "").strip(): r for r in base_rows}

    # Section 1: volume_num 필터 적용 9건
    volume_filtered: list[tuple[dict, dict]] = []
    for r in p3_rows:
        q = (r.get("테스트용 질문") or "").strip()
        meta = extract_query_metadata(q)
        if "volume_num" in meta:
            br = base_by_q.get(q)
            if br:
                volume_filtered.append((br, r))

    # Section 2: 메타데이터 미적용 L1~L5 각 1건 control
    rng = random.Random(args.seed)
    by_level: dict[str, list[tuple[dict, dict]]] = {"L1": [], "L2": [], "L3": [], "L4": [], "L5": []}
    for r in p3_rows:
        q = (r.get("테스트용 질문") or "").strip()
        meta = extract_query_metadata(q)
        if "volume_num" in meta:
            continue  # 이미 Section 1
        L = detect_level(str(r.get("난이도(Level)", "")))
        if L in by_level:
            br = base_by_q.get(q)
            if br:
                by_level[L].append((br, r))

    control: list[tuple[dict, dict]] = []
    for L in ("L1", "L2", "L3", "L4", "L5"):
        bucket = by_level.get(L, [])
        if bucket:
            control.append(rng.choice(bucket))

    # 마크다운 생성
    lines = [
        "# Codex 독립 정성 검토 — Phase 3 메타데이터 필터 (v5 baseline vs v5+메타필터)",
        "",
        "## 평가 요청",
        "",
        f"아래 {len(volume_filtered)}건(권번호 필터 적용) + {len(control)}건(필터 미적용 control) 비교 사례에서, ",
        "**v5 baseline**과 **Phase 3 (v5+메타필터)** 중 어느 답변이 더 정확하고 유용한지 판정해 주세요.",
        "",
        "각 사례마다:",
        "- **승자**: baseline / Phase 3 / 동등",
        "- **이유**: 모범답변/참고키워드 대비 정확도, 컨텍스트 활용도, 환각 여부 기준",
        "",
        "마지막에 종합 의견:",
        "- Section 1 (필터 적용 9건): Phase 3 우월 N건 / baseline 우월 M건 / 동등 K건",
        "- Section 2 (control 5건): 부작용 여부 (필터 미적용 케이스에서 변화)",
        "- **운영 적용 권장 여부** + 보강할 약점",
        "- 후속 PR 우선순위 (책별 분리 / 점수 부스팅 / date filter / Reranker 가중치)",
        "",
        "**참고 — 자동 메트릭 결과**:",
        "- 전체 LLM-Judge: 18.03 (동률, ±0)",
        "- L2 LLM-Judge: 17.08 → 17.55 (+0.47)",
        "- 필터 9건 평균 LLM-Judge: +0.78점, 015번 +7.5점",
        "- 채택 임계값(LLM-Judge ≥ 18.5) 미달 → 운영 적용 보류",
        "",
        "---",
        "",
        "# Section 1 — volume_num 필터 적용 9건 (메타필터 직접 영향)",
        "",
    ]

    for i, (br, pr) in enumerate(volume_filtered, 1):
        b_a, b_c = extract_pair(br, base_ans)
        p_a, p_c = extract_pair(pr, p3_ans)
        lines.extend(render_case(i, pr, b_a, b_c, p_a, p_c, "S1"))

    lines.append("")
    lines.append("# Section 2 — 메타필터 미적용 control 5건 (부작용 검증)")
    lines.append("")
    for i, (br, pr) in enumerate(control, 1):
        b_a, b_c = extract_pair(br, base_ans)
        p_a, p_c = extract_pair(pr, p3_ans)
        lines.extend(render_case(i, pr, b_a, b_c, p_a, p_c, "S2"))

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text("\n".join(lines), encoding="utf-8")
    total = len(volume_filtered) + len(control)
    print(f"✓ Codex 비교 markdown {total}건 생성: {args.output}")
    print(f"  Section 1 (filter 적용): {len(volume_filtered)}건")
    print(f"  Section 2 (control): {len(control)}건")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
