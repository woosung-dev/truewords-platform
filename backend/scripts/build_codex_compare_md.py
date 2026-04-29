"""A/F 또는 A/B/F 측정 xlsx 비교 Codex 검토용 마크다운 생성.

L별 2건씩 stratified 10건 추출 → Codex가 비교하기 좋은 형식.

2-way 사용 (A vs F):
    PYTHONPATH=. uv run python scripts/build_codex_compare_md.py \\
        --xlsx-a ~/Downloads/notebooklm_qa_A_sentence_new50_*.xlsx \\
        --xlsx-f ~/Downloads/notebooklm_qa_F_paragraph_new50_*.xlsx \\
        --output ~/Downloads/codex_compare_input.md

3-way 사용 (A vs B vs F):
    PYTHONPATH=. uv run python scripts/build_codex_compare_md.py \\
        --xlsx-a ~/Downloads/notebooklm_qa_A_*.xlsx \\
        --xlsx-b ~/Downloads/notebooklm_qa_B_*.xlsx \\
        --xlsx-f ~/Downloads/notebooklm_qa_F_*.xlsx \\
        --output ~/Downloads/codex_compare_3way.md
"""
from __future__ import annotations

import argparse
import random
import sys
from pathlib import Path

from openpyxl import load_workbook


def detect_level(level_raw: str) -> str:
    for L in ("L1", "L2", "L3", "L4", "L5"):
        if L in level_raw:
            return L
    return "기타"


def find_answer_col(headers: list[str]) -> str | None:
    for h in headers:
        if "우리 답변" in h:
            return h
    return None


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        out.append({headers[i]: r[i] if i < len(r) else None for i in range(len(headers))})
    return out


def _extract_answer_and_ctx(row: dict | None, ans_col: str | None) -> tuple[str, list[str]]:
    if row is None:
        return "(매칭 실패)", []
    answer = str(row.get(ans_col, "") if ans_col else "")
    ctx = []
    for col in ("참고1", "참고2", "참고3"):
        c = row.get(col)
        if c:
            ctx.append(str(c)[:300])
    return answer, ctx


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-a", required=True, type=Path)
    parser.add_argument("--xlsx-f", required=True, type=Path)
    parser.add_argument("--xlsx-b", type=Path, default=None,
                        help="옵션 B (prefix) — 있으면 3-way 비교")
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--per-level", type=int, default=2)
    parser.add_argument("--seed", type=int, default=20260429)
    args = parser.parse_args()

    rows_a = load_rows(args.xlsx_a)
    rows_f = load_rows(args.xlsx_f)
    if not rows_a or not rows_f:
        print(f"⚠️  로드 실패")
        return 1

    has_b = args.xlsx_b is not None
    rows_b = load_rows(args.xlsx_b) if has_b else []
    if has_b and not rows_b:
        print(f"⚠️  B xlsx 로드 실패: {args.xlsx_b}")
        return 1

    ans_a = find_answer_col(list(rows_a[0].keys()))
    ans_f = find_answer_col(list(rows_f[0].keys()))
    ans_b = find_answer_col(list(rows_b[0].keys())) if has_b else None

    by_q_f = {(r.get("테스트용 질문") or "").strip(): r for r in rows_f}
    by_q_b = {(r.get("테스트용 질문") or "").strip(): r for r in rows_b} if has_b else {}

    # L별 stratify (A를 기준)
    rng = random.Random(args.seed)
    by_level: dict[str, list[dict]] = {}
    for r in rows_a:
        L = detect_level(str(r.get("난이도(Level)", "")))
        if L not in ("L1", "L2", "L3", "L4", "L5"):
            continue
        by_level.setdefault(L, []).append(r)

    selected: list[dict] = []
    for L in ("L1", "L2", "L3", "L4", "L5"):
        bucket = by_level.get(L, [])
        n = min(args.per_level, len(bucket))
        picks = rng.sample(bucket, n) if n else []
        selected.extend(picks)

    # 마크다운 생성
    lines = []
    if has_b:
        lines.append("# Codex 독립 검토 — 옵션 A (sentence) vs B (prefix) vs F (paragraph)")
        lines.append("")
        lines.append("## 평가 요청")
        lines.append("")
        lines.append("아래 10건의 비교 사례에서 **A, B, F 중 어느 답변이 가장 정확하고 유용한가**를")
        lines.append("판정해 주세요. 각 사례마다:")
        lines.append("")
        lines.append("- **승자**: A / B / F / 동등 (또는 부분 동등 명시)")
        lines.append("- **이유**: 모범답변/참고키워드 대비 정확도, 컨텍스트 활용도, 환각 여부 기준")
        lines.append("")
        lines.append("마지막에 종합 의견:")
        lines.append("- 10건 중 A 승: N, B 승: M, F 승: K, 동등: L")
        lines.append("- 메트릭별(정확도/문맥/환각) 우열 패턴")
        lines.append("- L별(L1~L5) 패턴")
        lines.append("- 운영 적용 시 추천 옵션 + 보강할 약점")
    else:
        lines.append("# Codex 독립 검토 — 옵션 A (sentence) vs 옵션 F (paragraph)")
        lines.append("")
        lines.append("## 평가 요청")
        lines.append("")
        lines.append("아래 10건의 비교 사례에서 **A와 F 중 어느 답변이 더 정확하고 유용한가**를")
        lines.append("판정해 주세요. 각 사례마다:")
        lines.append("")
        lines.append("- **승자**: A / F / 동등")
        lines.append("- **이유**: 모범답변/참고키워드 대비 정확도, 컨텍스트 활용도, 환각 여부 기준")
        lines.append("")
        lines.append("마지막에 종합 의견:")
        lines.append("- 10건 중 A 승: N, F 승: M, 동등: K")
        lines.append("- 메트릭별(정확도/문맥/환각) 우열 패턴")
        lines.append("- 운영 적용 시 추천 옵션 + 보강할 약점")
    lines.append("")
    lines.append("---")
    lines.append("")

    for idx, r_a in enumerate(selected, 1):
        q = (r_a.get("테스트용 질문") or "").strip()
        L = detect_level(str(r_a.get("난이도(Level)", "")))
        cat = str(r_a.get("카테고리", "") or "")
        gt = str(r_a.get("봇 모범 답변", "") or "")
        kw = str(r_a.get("참고 키워드", "") or "")

        a_answer, a_ctx = _extract_answer_and_ctx(r_a, ans_a)
        f_answer, f_ctx = _extract_answer_and_ctx(by_q_f.get(q), ans_f)
        b_answer, b_ctx = (_extract_answer_and_ctx(by_q_b.get(q), ans_b) if has_b else ("", []))

        lines.append(f"## 사례 {idx} — {L} / {cat}")
        lines.append("")
        lines.append(f"**질문**: {q}")
        lines.append("")
        lines.append(f"**모범답변**: {gt}")
        lines.append("")
        lines.append(f"**참고키워드**: {kw}")
        lines.append("")

        for label, answer, ctx in [
            ("A (sentence chunking)", a_answer, a_ctx),
            *( [("B (prefix chunking)", b_answer, b_ctx)] if has_b else [] ),
            ("F (paragraph chunking)", f_answer, f_ctx),
        ]:
            lines.append(f"### 옵션 {label}")
            lines.append("")
            lines.append(f"**답변**: {answer[:1200]}")
            lines.append("")
            if ctx:
                lines.append("**컨텍스트 요약**:")
                for i, c in enumerate(ctx, 1):
                    lines.append(f"{i}. {c}")
                lines.append("")
        lines.append("---")
        lines.append("")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text("\n".join(lines), encoding="utf-8")
    mode = "3-way (A/B/F)" if has_b else "2-way (A/F)"
    print(f"생성: {args.output} ({len(selected)} 사례, {mode})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
