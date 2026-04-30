"""RAGAS 5건 batch 직렬 호출 (handoff 권고 패턴, 50건 hang 회피).

seed JSON을 batch_size씩 분할 → eval_ragas.py 반복 호출 → 결과 xlsx 합치기.

사용:
    PYTHONPATH=. uv run --group eval python scripts/eval_ragas_batched.py \\
        --seed ~/Downloads/ragas_seed_strat50_v1_*.json \\
        --output ~/Downloads/ragas_batched_v1.xlsx \\
        --batch-size 5
"""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
import tempfile
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--seed", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--batch-size", type=int, default=5)
    args = parser.parse_args()

    items = json.loads(args.seed.read_text())
    print(f"[load] {args.seed.name}: {len(items)}건, batch_size={args.batch_size}")

    n_batches = (len(items) + args.batch_size - 1) // args.batch_size
    batch_xlsxs: list[Path] = []
    t0 = time.time()

    for i in range(n_batches):
        batch_items = items[i * args.batch_size : (i + 1) * args.batch_size]
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".json", delete=False, encoding="utf-8"
        ) as tmp:
            json.dump(batch_items, tmp, ensure_ascii=False)
            tmp_seed = Path(tmp.name)

        batch_out = args.output.with_suffix("").with_name(
            f"{args.output.stem}_batch{i:02d}.xlsx"
        )
        batch_xlsxs.append(batch_out)

        elapsed = time.time() - t0
        print(f"[{i+1}/{n_batches}] batch {len(batch_items)}건 ({elapsed/60:.1f}분 경과) ...")

        # eval_ragas.py 호출
        result = subprocess.run(
            [
                "uv", "run", "--group", "eval", "python", "-u",
                "scripts/eval_ragas.py",
                "--seed", str(tmp_seed),
                "--output", str(batch_out),
            ],
            capture_output=True,
            text=True,
            timeout=600,  # 10분 타임아웃 (5건당)
            cwd=Path(__file__).resolve().parent.parent,
        )
        if result.returncode != 0:
            print(f"  ⚠️  batch {i+1} 실패: {result.stderr[-500:]}")
            tmp_seed.unlink(missing_ok=True)
            continue

        # 결과 line 추출 (mean=...)
        for line in result.stdout.splitlines():
            if "mean=" in line:
                print(f"  {line.strip()}")
        tmp_seed.unlink(missing_ok=True)

    # 결과 xlsx 합치기 — 첫 batch xlsx의 시트 구조로 통합
    print(f"\n[merge] {len(batch_xlsxs)} batch xlsx → {args.output.name}")
    merged_wb = Workbook()
    merged_ws = merged_wb.active
    if merged_ws is None:
        return 1
    merged_ws.title = "RAGAS 결과"

    header_written = False
    summary_metrics: dict[str, list[float]] = {}

    for batch_xlsx in batch_xlsxs:
        if not batch_xlsx.exists():
            continue
        wb = load_workbook(batch_xlsx, data_only=True)
        # 첫 시트(상세 결과) 사용
        for ws in wb.worksheets:
            if "결과" in ws.title or ws.title == wb.sheetnames[0]:
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                if not header_written:
                    merged_ws.append(list(rows[0]))
                    for cell in merged_ws[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill("solid", fgColor="D9E1F2")
                    header_written = True
                # 데이터 row만
                for row in rows[1:]:
                    if all(c is None or c == "" for c in row):
                        continue
                    merged_ws.append(list(row))
                    # 메트릭 컬럼 (faithfulness/context_precision/context_recall/answer_relevancy)
                    headers = list(rows[0])
                    for col_name in ["faithfulness", "context_precision", "context_recall", "answer_relevancy"]:
                        if col_name in headers:
                            idx = headers.index(col_name)
                            if idx < len(row):
                                v = row[idx]
                                if isinstance(v, (int, float)):
                                    summary_metrics.setdefault(col_name, []).append(float(v))
                break

    # 요약 시트
    summary = merged_wb.create_sheet("요약")
    summary.append(["메트릭", "평균", "n"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for metric, values in summary_metrics.items():
        if values:
            avg = sum(values) / len(values)
            summary.append([metric, f"{avg:.3f}", len(values)])
            print(f"  {metric:<25s} mean={avg:.3f}  n={len(values)}")

    args.output.parent.mkdir(parents=True, exist_ok=True)
    merged_wb.save(args.output)

    # batch xlsx 정리
    for batch_xlsx in batch_xlsxs:
        batch_xlsx.unlink(missing_ok=True)

    elapsed = time.time() - t0
    print(f"\n[done] {args.output} ({elapsed/60:.1f}분 소요)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
