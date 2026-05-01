"""RAGAS 시드 50건을 원본 xlsx와 join해서 gold_source(가장 자주 등장한 source 코드)를 부여.

시드의 contexts는 본문만 보유 (sample_eval_pairs.py가 marker 제거)이므로, 원본 xlsx의
참고1/참고2/참고3 셀에서 ``[문서명] (score=X, source=Y)`` 패턴을 정규식으로 추출해 join.

시드 id 형식: ``<basename>#<번호>`` → xlsx의 ``번호`` 컬럼과 join.

사용 예:
    PYTHONPATH=. uv run python scripts/label_seed_with_source.py \\
        --input ~/Downloads/ragas_eval_seed_50_20260427_2306.json \\
        --output ~/Downloads/ragas_eval_seed_50_with_source_label.json \\
        --xlsx-dir ~/Downloads
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

SOURCE_RE = re.compile(r"source=([A-T])")


def _build_source_index(xlsx_path: Path) -> dict[str, list[str]]:
    """xlsx 한 파일에서 ``번호 -> [source codes from 참고1+2+3]`` 매핑."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    if ws is None:
        return {}
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(c) if c else "" for c in rows[0]]
    idx: dict[str, list[str]] = {}
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        d = dict(zip(header, r))
        no = str(d.get("번호") or "").strip()
        if not no:
            continue
        joined = " ".join(str(d.get(k) or "") for k in ("참고1", "참고2", "참고3"))
        idx[no] = SOURCE_RE.findall(joined)
    return idx


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--input", required=True, type=Path)
    p.add_argument("--output", required=True, type=Path)
    p.add_argument("--xlsx-dir", required=True, type=Path)
    args = p.parse_args()
    seed = json.loads(args.input.read_text(encoding="utf-8"))

    # source_file 별로 한 번만 xlsx 인덱스 빌드
    cache: dict[str, dict[str, list[str]]] = {}

    out: list[dict] = []
    unknown = 0
    for row in seed:
        sf = row.get("source_file") or ""
        if sf and sf not in cache:
            xlsx = args.xlsx_dir / sf
            cache[sf] = _build_source_index(xlsx) if xlsx.exists() else {}
        idx = cache.get(sf, {})
        # id 형식: <basename>#<번호>
        rid = str(row.get("id") or "")
        no = rid.split("#")[-1] if "#" in rid else ""
        codes = idx.get(no, [])
        counts = Counter(codes)
        gold = counts.most_common(1)[0][0] if counts else "Unknown"
        secondary = counts.most_common(2)[1][0] if len(counts) > 1 else None
        if gold == "Unknown":
            unknown += 1
        out.append({**row, "gold_source": gold, "gold_source_secondary": secondary})

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(out, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"라벨링 완료: {args.output} | {len(out)}건 (Unknown {unknown}건)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
