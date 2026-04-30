"""기존 측정 xlsx (v1/v2/v3) → 동일 50문항 stratified sample → 3 seed JSON 생성.

목적: v1/v2/v3 RAGAS 4메트릭 비교 (apples-to-apples).

각 xlsx 컬럼 (eval_notebooklm_qa.py 출력 포맷):
    번호, 난이도(Level), 카테고리, 테스트용 질문, 봇 모범 답변, 참고 키워드,
    우리 답변(<botname>), 참고1, 참고2, 참고3, 세션ID

출력: 3 seed JSON (eval_ragas.py 호환)
    [{"id", "source_file", "level", "category", "question", "answer",
      "ground_truth", "contexts": list[str]}, ...]

사용:
    PYTHONPATH=. uv run python scripts/build_ragas_seeds_from_ab.py \\
        --output-dir ~/Downloads
"""
from __future__ import annotations

import argparse
import json
import random
import re
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


# 비교 대상 측정 xlsx (v3는 'all' 봇 토글 재측정 — 봇 설정 공평)
MEASUREMENT_FILES = {
    "v1": "/Users/woosung/Downloads/notebooklm_post_phase1_20260428_1001_light100.xlsx",
    "v2": "/Users/woosung/Downloads/notebooklm_v2_20260428_2338_light100.xlsx",
    "v3new": "/Users/woosung/Downloads/notebooklm_v3_remeasure_20260429_1716.xlsx",
}

# Stratified sample: 100건 전체 (L1~L5 각 ~20건). 사용자 요청: 꼼꼼하게.
LEVEL_QUOTA = {"L1": 100, "L2": 100, "L3": 100, "L4": 100, "L5": 100}  # 전체 사용


CONTEXT_PATTERN = re.compile(
    r"^\[[^\]]+\]\s*\([^)]+\)\s*\n(.+)$",
    re.DOTALL,
)


def parse_context_cell(cell_value: str | None) -> str | None:
    """참고1/2/3 셀에서 본문 텍스트만 추출. '[문서명] (score=..., source=...)\\n<본문>' 형식."""
    if not cell_value or not isinstance(cell_value, str):
        return None
    m = CONTEXT_PATTERN.match(cell_value.strip())
    if m:
        return m.group(1).strip()
    return cell_value.strip()


def load_rows(path: Path) -> list[dict]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for r in rows[1:]:
        if not r or all(c is None or c == "" for c in r):
            continue
        d = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
        out.append(d)
    return out


def find_answer_col(headers: list[str]) -> str | None:
    for h in headers:
        if "우리 답변" in h:
            return h
    return None


def detect_level(level_raw: str) -> str:
    for L in ("L1", "L2", "L3", "L4", "L5"):
        if L in level_raw:
            return L
    return "기타"


def build_seed_single(xlsx_path: Path, label: str, output_dir: Path, ts: str) -> Path | None:
    """단일 측정 xlsx → seed JSON. stratify 없이 전체 사용. keywords 컬럼 포함."""
    rows = load_rows(xlsx_path)
    if not rows:
        print(f"⚠️  {label} 로드 실패: {xlsx_path}")
        return None
    headers = list(rows[0].keys())
    ans_col = find_answer_col(headers)

    seed_items = []
    for idx, r in enumerate(rows, 1):
        q = (r.get("테스트용 질문") or "").strip()
        if not q:
            continue
        answer = str(r.get(ans_col, "") if ans_col else "").strip()
        ground_truth = str(r.get("봇 모범 답변", "")).strip()
        keywords_raw = str(r.get("참고 키워드", "") or "").strip()
        keywords = [k.strip() for k in keywords_raw.split(",") if k.strip()]
        contexts = []
        for col in ("참고1", "참고2", "참고3"):
            c = parse_context_cell(r.get(col))
            if c:
                contexts.append(c)
        level = detect_level(str(r.get("난이도(Level)", "")))
        seed_items.append({
            "id": f"{label}_{idx:03d}",
            "source_file": xlsx_path.name,
            "level": level,
            "category": str(r.get("카테고리", "") or ""),
            "question": q,
            "answer": answer,
            "ground_truth": ground_truth,
            "keywords": keywords,
            "contexts": contexts,
        })

    seed_path = output_dir / f"ragas_seed_{label}_{ts}.json"
    seed_path.write_text(json.dumps(seed_items, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  {label}: {len(seed_items)}건 → {seed_path.name}")
    return seed_path


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", default=Path.home() / "Downloads", type=Path)
    parser.add_argument("--seed", type=int, default=20260429)
    parser.add_argument("--xlsx", type=Path, default=None,
                        help="단일 파일 모드: 측정 xlsx 경로")
    parser.add_argument("--label", type=str, default=None,
                        help="단일 파일 모드: seed JSON 라벨 (예: A_new50, F_new50)")
    args = parser.parse_args()

    # 단일 파일 모드 (--xlsx + --label)
    if args.xlsx is not None:
        if args.label is None:
            print("⚠️  --xlsx와 --label은 함께 지정해야 합니다.")
            return 1
        args.output_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        seed_path = build_seed_single(args.xlsx, args.label, args.output_dir, ts)
        if seed_path is None:
            return 1
        print()
        print("=== 다음 단계 ===")
        out = args.output_dir / f"ragas_{args.label}_{ts}.xlsx"
        print(f"  RAGAS: PYTHONPATH=. uv run python scripts/eval_metrics_direct.py "
              f"--seed {seed_path} --output {out}")
        return 0

    # 기존 다중 파일 모드 (legacy v1/v2/v3)
    # v1을 기준으로 50문항 sample 선택 (질문 ID = "테스트용 질문" 텍스트)
    rng = random.Random(args.seed)

    v1_path = Path(MEASUREMENT_FILES["v1"])
    v1_rows = load_rows(v1_path)
    print(f"v1 로드: {len(v1_rows)}건 ({v1_path.name})")

    # level별 grouping
    by_level: dict[str, list[dict]] = {L: [] for L in LEVEL_QUOTA.keys()}
    for r in v1_rows:
        L = detect_level(str(r.get("난이도(Level)", "")))
        if L in by_level:
            by_level[L].append(r)

    sample_questions: list[str] = []
    for L, quota in LEVEL_QUOTA.items():
        bucket = by_level.get(L, [])
        if len(bucket) < quota:
            print(f"  ⚠️  L{L} bucket < quota ({len(bucket)}/{quota}) — 전체 사용")
            picked = bucket
        else:
            picked = rng.sample(bucket, quota)
        for r in picked:
            q = (r.get("테스트용 질문") or "").strip()
            if q:
                sample_questions.append(q)
        print(f"  {L}: {len(picked)}건 sample")

    print(f"총 sample: {len(sample_questions)} questions")

    # 각 측정 xlsx에서 동일 question으로 답변+contexts 추출 → seed JSON 생성
    args.output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    seed_paths: dict[str, Path] = {}

    for ver, path_str in MEASUREMENT_FILES.items():
        path = Path(path_str)
        rows = load_rows(path)
        rows_by_q = {(r.get("테스트용 질문") or "").strip(): r for r in rows}
        if not rows:
            print(f"⚠️  {ver} 로드 실패: {path}")
            continue

        headers = list(rows[0].keys())
        ans_col = find_answer_col(headers)

        seed_items = []
        for idx, q in enumerate(sample_questions, 1):
            r = rows_by_q.get(q)
            if not r:
                print(f"  ⚠️  {ver}: '{q[:30]}...' 매칭 안 됨")
                continue
            answer = str(r.get(ans_col, "") if ans_col else "").strip()
            ground_truth = str(r.get("봇 모범 답변", "")).strip()
            contexts = []
            for col in ("참고1", "참고2", "참고3"):
                c = parse_context_cell(r.get(col))
                if c:
                    contexts.append(c)
            level = detect_level(str(r.get("난이도(Level)", "")))
            seed_items.append({
                "id": f"{ver}_{idx:03d}",
                "source_file": path.name,
                "level": level,
                "category": str(r.get("카테고리", "") or ""),
                "question": q,
                "answer": answer,
                "ground_truth": ground_truth,
                "contexts": contexts,
            })

        seed_path = args.output_dir / f"ragas_seed_50_{ver}_{ts}.json"
        seed_path.write_text(json.dumps(seed_items, ensure_ascii=False, indent=2), encoding="utf-8")
        seed_paths[ver] = seed_path
        print(f"  {ver}: {len(seed_items)}건 → {seed_path.name}")

    print()
    print("=== 다음 단계 ===")
    for ver, p in seed_paths.items():
        out = args.output_dir / f"ragas_50_{ver}_{ts}.xlsx"
        print(f"  {ver}: PYTHONPATH=. uv run --group eval python scripts/eval_ragas.py "
              f"--seed {p} --output {out} --limit 5")
    return 0


if __name__ == "__main__":
    sys.exit(main())
