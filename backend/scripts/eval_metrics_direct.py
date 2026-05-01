"""LangChain 우회 — Gemini SDK 직접 호출로 4메트릭 평가.

ragas 0.4.3 + langchain-google-genai 4.2.2 환경 hang 회피용.
Gemini API 직접 호출이라 안정 동작 보장.

4메트릭 (RAGAS 의미 그대로):
- faithfulness: 답변이 컨텍스트에 충실한가 (hallucination 검출)
- context_precision: 컨텍스트가 정답에 정렬되는가
- context_recall: 정답이 컨텍스트에 회수되는가
- response_relevancy: 답변이 질문에 관련 있는가

사용:
    PYTHONPATH=. uv run python scripts/eval_metrics_direct.py \\
        --seed ~/Downloads/ragas_seed_strat50_v1_*.json \\
        --output ~/Downloads/metrics_direct_v1.xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import json
import re
import sys
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from google import genai
from google.genai import types

from src.config import settings


EVAL_MODEL = "gemini-3.1-flash-lite-preview"  # RPD 150K (handoff 메모리)
CONCURRENCY = 5


EVAL_PROMPT = """다음 RAG 시스템의 응답을 4가지 메트릭으로 평가하세요.

[질문]
{question}

[답변]
{answer}

[검색된 컨텍스트]
{contexts}

[정답 (ground_truth)]
{ground_truth}

평가 기준 (0.0~1.0, RAGAS 메트릭 정의 그대로):

1. **faithfulness**: 답변의 각 주장이 컨텍스트에 의해 지원되는 비율.
   답변에서 사실 주장을 분해 → 컨텍스트로 입증 가능한 비율. (hallucination 검출)
   1.0 = 모든 주장이 컨텍스트로 입증됨, 0.0 = 모든 주장이 hallucination.

2. **context_precision**: 검색된 컨텍스트 중 정답과 관련된 청크의 비율.
   각 컨텍스트가 정답 도출에 직접 도움 되는지 평가.
   1.0 = 모든 컨텍스트가 관련, 0.0 = 모두 무관.

3. **context_recall**: 정답의 각 주장이 컨텍스트에 의해 회수 가능한 비율.
   정답에서 주장을 분해 → 컨텍스트에 그 주장이 있는지 확인.
   1.0 = 정답의 모든 주장이 컨텍스트에 있음, 0.0 = 전혀 없음.

4. **response_relevancy**: 답변이 질문에 직접 관련된 정도.
   답변이 질문 의도에 정확히 답하는지. 부적절한 정보 포함은 감점.
   1.0 = 완벽 관련, 0.0 = 무관 또는 회피.

JSON으로만 출력 (다른 텍스트 금지):
{{"faithfulness": 0.XX, "context_precision": 0.XX, "context_recall": 0.XX, "response_relevancy": 0.XX, "reasoning": "각 메트릭 점수 근거 한 줄씩 (총 4줄)"}}
"""


def _parse_score(text: str) -> dict | None:
    """LLM 응답에서 JSON 추출."""
    # ```json ... ``` 또는 raw JSON
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        text = m.group(1)
    else:
        m = re.search(r"(\{[\s\S]*?\})", text)
        if m:
            text = m.group(1)
    try:
        d = json.loads(text)
        return {
            "faithfulness": float(d.get("faithfulness", 0)),
            "context_precision": float(d.get("context_precision", 0)),
            "context_recall": float(d.get("context_recall", 0)),
            "response_relevancy": float(d.get("response_relevancy", 0)),
            "reasoning": str(d.get("reasoning", "")),
        }
    except (json.JSONDecodeError, ValueError, TypeError):
        return None


async def evaluate_one(
    client, item: dict, semaphore: asyncio.Semaphore, idx: int, total: int
) -> dict:
    """단일 sample 4메트릭 평가."""
    contexts_text = "\n\n".join(
        f"[컨텍스트 {i+1}]\n{c}" for i, c in enumerate(item.get("contexts", []))
    )
    prompt = EVAL_PROMPT.format(
        question=item.get("question", ""),
        answer=item.get("answer", ""),
        contexts=contexts_text,
        ground_truth=item.get("ground_truth", ""),
    )

    async with semaphore:
        for attempt in range(3):
            try:
                resp = await client.aio.models.generate_content(
                    model=EVAL_MODEL,
                    contents=prompt,
                    config=types.GenerateContentConfig(temperature=0),
                )
                scores = _parse_score(resp.text or "")
                if scores:
                    print(
                        f"  [{idx+1}/{total}] f={scores['faithfulness']:.2f} "
                        f"cp={scores['context_precision']:.2f} "
                        f"cr={scores['context_recall']:.2f} "
                        f"ar={scores['response_relevancy']:.2f}",
                        flush=True,
                    )
                    return {
                        "id": item.get("id", str(idx)),
                        "level": item.get("level", ""),
                        "question": item.get("question", ""),
                        **scores,
                    }
                print(f"  [{idx+1}/{total}] parse 실패, retry {attempt+1}/3", flush=True)
            except Exception as e:
                print(
                    f"  [{idx+1}/{total}] error: {type(e).__name__}: {e} (retry {attempt+1}/3)",
                    flush=True,
                )
                await asyncio.sleep(5)
    # 3회 실패
    return {
        "id": item.get("id", str(idx)),
        "level": item.get("level", ""),
        "question": item.get("question", ""),
        "faithfulness": None,
        "context_precision": None,
        "context_recall": None,
        "response_relevancy": None,
        "reasoning": "FAILED",
    }


async def main_async(args) -> int:
    items = json.loads(args.seed.read_text(encoding="utf-8"))
    if args.limit:
        items = items[: args.limit]
    print(f"[load] {args.seed.name}: {len(items)}건 (concurrency={CONCURRENCY})")

    client = genai.Client(api_key=settings.gemini_api_key.get_secret_value())
    semaphore = asyncio.Semaphore(CONCURRENCY)

    t0 = time.time()
    tasks = [
        evaluate_one(client, item, semaphore, i, len(items))
        for i, item in enumerate(items)
    ]
    results = await asyncio.gather(*tasks)
    elapsed = time.time() - t0

    # 평균 계산 (None 제외)
    metrics = ["faithfulness", "context_precision", "context_recall", "response_relevancy"]
    avgs: dict[str, float | None] = {}
    for m in metrics:
        values = [r[m] for r in results if r[m] is not None]
        avgs[m] = sum(values) / len(values) if values else None

    # xlsx 저장
    wb = Workbook()
    ws = wb.active
    if ws is None:
        return 1
    ws.title = "metrics"
    headers = [
        "id", "level", "question",
        "faithfulness", "context_precision", "context_recall", "response_relevancy",
        "reasoning",
    ]
    ws.append(headers)
    bold = Font(bold=True)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    for r in results:
        ws.append([
            r["id"], r["level"], r["question"][:200],
            r["faithfulness"], r["context_precision"], r["context_recall"], r["response_relevancy"],
            r.get("reasoning", "")[:500],
        ])

    # 요약 시트
    summary = wb.create_sheet("요약")
    summary.append(["메트릭", "평균", "n"])
    for cell in summary[1]:
        cell.font = bold
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for m in metrics:
        v = avgs[m]
        n = sum(1 for r in results if r[m] is not None)
        summary.append([m, f"{v:.3f}" if v is not None else "-", n])

    args.output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.output)
    print()
    print(f"=== 결과 ({elapsed:.1f}s) ===")
    for m in metrics:
        v = avgs[m]
        n = sum(1 for r in results if r[m] is not None)
        print(f"  {m:<25s} mean={v:.3f}  n={n}/{len(items)}" if v is not None else f"  {m:<25s} 측정 실패")
    print(f"저장: {args.output}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--seed", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--limit", type=int, default=None)
    args = parser.parse_args()
    return asyncio.run(main_async(args))


if __name__ == "__main__":
    sys.exit(main())
