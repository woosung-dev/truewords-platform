"""TrueWords 운영 DB 실제 대화 전체 추출 스크립트.

운영 Neon PostgreSQL의 session_messages / search_events / answer_citations /
answer_feedback / chatbot_configs 를 LATERAL JOIN 으로 묶어서 한 행 = (질문, 답변,
출처, 피드백, 메타) 형태의 NotebookLM 스타일 .xlsx 로 저장한다.

동일 질문 정책: 정규화(trim + lowercase) 후 같은 질문은 **가장 최근 답변**만 유지
(`DISTINCT ON ... ORDER BY ... answered_at DESC`).

⚠️ 자격증명: DATABASE_URL 은 환경변수로만 받는다. 코드/파일에 적지 않는다.
  사용 예:
    export DATABASE_URL=$(gcloud run services describe truewords-backend \\
      --region asia-northeast3 --format=json | jq -r \\
      '.spec.template.spec.containers[0].env[] | select(.name=="DATABASE_URL") | .value')
    uv run --with asyncpg --with openpyxl python docs/guides/scripts/fetch_real_conversations.py

산출물:
  docs/guides/redteam-prod-real-conversations.xlsx
  docs/guides/redteam-prod-real-conversations.summary.json
"""
from __future__ import annotations

import asyncio
import json
import os
import re
import sys
import time
from pathlib import Path

import asyncpg
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_XLSX = Path(__file__).resolve().parents[1] / "redteam-prod-real-conversations.xlsx"
OUT_SUMMARY = Path(__file__).resolve().parents[1] / "redteam-prod-real-conversations.summary.json"

# 운영 DB 에서 user-assistant 페어 + 검색/출처/피드백/챗봇 메타를 한 번에 가져옴.
# DISTINCT ON 으로 동일 질문 중 가장 최근 답변만 유지.
SQL = """
WITH paired AS (
  SELECT
    s.id            AS session_id,
    s.started_at    AS session_started_at,
    cc.chatbot_id   AS chatbot_key,
    cc.display_name AS chatbot_name,
    um.id           AS user_msg_id,
    um.content      AS user_question,
    um.created_at   AS asked_at,
    am.id           AS assistant_msg_id,
    am.content      AS assistant_answer,
    am.created_at   AS answered_at,
    am.resolved_answer_mode,
    am.persona_overridden,
    am.crisis_trigger,
    se.query_text       AS search_query,
    se.rewritten_query  AS rewritten_query,
    se.search_tier,
    se.total_results,
    se.latency_ms,
    cit.sources_text,
    fb.feedbacks_text,
    EXTRACT(EPOCH FROM (am.created_at - um.created_at))::float AS response_seconds
  FROM session_messages um
  JOIN research_sessions s ON s.id = um.session_id
  LEFT JOIN chatbot_configs cc ON cc.id = s.chatbot_config_id
  JOIN LATERAL (
    SELECT m2.* FROM session_messages m2
    WHERE m2.session_id = um.session_id
      AND m2.role = 'ASSISTANT'
      AND m2.created_at > um.created_at
    ORDER BY m2.created_at ASC
    LIMIT 1
  ) am ON TRUE
  -- search_events.message_id 는 ASSISTANT 메시지 id 를 가리킴 (운영 검증).
  LEFT JOIN search_events se ON se.message_id = am.id
  LEFT JOIN LATERAL (
    SELECT string_agg(
      format('[%s] %s (%s) · score=%s',
             rank_position + 1,
             COALESCE(volume_raw, volume::text),
             source,
             to_char(round(relevance_score::numeric, 2), 'FM0.00')),
      E'\\n' ORDER BY rank_position
    ) AS sources_text
    FROM answer_citations c
    WHERE c.message_id = am.id
  ) cit ON TRUE
  LEFT JOIN LATERAL (
    SELECT string_agg(
      format('%s%s', feedback_type::text,
             CASE WHEN comment IS NULL OR comment = '' THEN '' ELSE ': ' || comment END),
      E'\\n' ORDER BY created_at
    ) AS feedbacks_text
    FROM answer_feedback f
    WHERE f.message_id = am.id
  ) fb ON TRUE
  WHERE um.role = 'USER'
)
SELECT * FROM (
  SELECT DISTINCT ON (lower(btrim(user_question)))
    session_id, session_started_at, chatbot_key, chatbot_name,
    user_msg_id, user_question, asked_at,
    assistant_msg_id, assistant_answer, answered_at,
    resolved_answer_mode, persona_overridden, crisis_trigger,
    search_query, rewritten_query, search_tier, total_results, latency_ms,
    sources_text, feedbacks_text, response_seconds
  FROM paired
  ORDER BY lower(btrim(user_question)), answered_at DESC
) deduped
ORDER BY answered_at DESC;  -- 최종 행 정렬: 최신 답변이 가장 위
"""


async def fetch_rows() -> list[dict]:
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("ERROR: DATABASE_URL 환경변수가 비어있습니다.", file=sys.stderr)
        print("       gcloud 명령으로 export 후 다시 실행하세요 (스크립트 docstring 참조).", file=sys.stderr)
        sys.exit(2)

    # asyncpg 는 SQLAlchemy 의 +asyncpg 접두사를 받지 않음
    db_url = re.sub(r"^postgresql\+asyncpg://", "postgresql://", db_url)
    # asyncpg 는 sslmode 쿼리스트링 형식 인식하지 않음 — ssl=require 로 통일
    if "sslmode=require" in db_url and "ssl=require" not in db_url:
        db_url = db_url.replace("sslmode=require", "ssl=require")

    print(f"▶ Neon 연결 중…")
    conn = await asyncpg.connect(db_url, ssl="require" if "ssl=require" in db_url else None)
    try:
        t0 = time.monotonic()
        rows = await conn.fetch(SQL)
        elapsed = time.monotonic() - t0
        print(f"   조회 완료: {len(rows):,}개 페어 / {elapsed:.1f}s")
        return [dict(r) for r in rows]
    finally:
        await conn.close()


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    assert wb.active is not None
    ws = wb.active
    ws.title = "실제 대화 (dedup)"

    headers = [
        "#", "챗봇", "Chatbot ID", "질문", "답변",
        "출처 (volume / source / score)", "피드백",
        "답변 모드", "페르소나 override", "위기 트리거",
        "원본 검색 query", "재작성 query", "Tier", "검색 결과 수", "검색 지연(ms)",
        "응답 소요(s)",
        "asked_at (UTC)", "answered_at (UTC)",
        "session_id", "user_msg_id", "assistant_msg_id",
    ]
    ws.append(headers)

    # 답변 너무 길면 한 셀에 다 못 들어갈 수 있어 32,767자 제한 (Excel cell limit) 내로
    EXCEL_CELL_LIMIT = 32_700

    for i, r in enumerate(rows, 1):
        ws.append([
            i,
            r.get("chatbot_name") or "",
            r.get("chatbot_key") or "",
            (r.get("user_question") or "")[:EXCEL_CELL_LIMIT],
            (r.get("assistant_answer") or "")[:EXCEL_CELL_LIMIT],
            (r.get("sources_text") or "")[:EXCEL_CELL_LIMIT],
            (r.get("feedbacks_text") or "")[:EXCEL_CELL_LIMIT],
            r.get("resolved_answer_mode") or "",
            "" if r.get("persona_overridden") is None else ("Y" if r["persona_overridden"] else "N"),
            r.get("crisis_trigger") or "",
            (r.get("search_query") or "")[:EXCEL_CELL_LIMIT],
            (r.get("rewritten_query") or "")[:EXCEL_CELL_LIMIT],
            r.get("search_tier") if r.get("search_tier") is not None else "",
            r.get("total_results") if r.get("total_results") is not None else "",
            r.get("latency_ms") if r.get("latency_ms") is not None else "",
            None if r.get("response_seconds") is None else round(r["response_seconds"], 1),
            r["asked_at"].strftime("%Y-%m-%d %H:%M:%S") if r.get("asked_at") else "",
            r["answered_at"].strftime("%Y-%m-%d %H:%M:%S") if r.get("answered_at") else "",
            str(r.get("session_id") or ""),
            str(r.get("user_msg_id") or ""),
            str(r.get("assistant_msg_id") or ""),
        ])

    # 헤더 스타일
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    widths = [5, 18, 22, 38, 60, 32, 22, 14, 14, 16, 32, 32, 6, 10, 12, 11, 18, 18, 22, 22, 22]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32

    # 본문 wrap
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 챗봇별 행 색상 (자동 분배)
    palette = ["DBE5F1", "FFF2CC", "E2EFDA", "FCE4D6", "EAD1DC", "D9E1F2", "FFE699", "C6E0B4"]
    bot_color: dict[str, str] = {}
    for row_idx in range(2, ws.max_row + 1):
        bot = ws.cell(row=row_idx, column=2).value or "?"
        if bot not in bot_color:
            bot_color[bot] = palette[len(bot_color) % len(palette)]
        color = bot_color[bot]
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid"
            )

    ws.freeze_panes = "E2"

    # Sheet 2: 요약 통계
    ws_meta = wb.create_sheet("요약 통계")
    total = len(rows)
    by_bot: dict[str, int] = {}
    feedback_count = 0
    persona_overridden_count = 0
    crisis_count = 0
    for r in rows:
        b = r.get("chatbot_name") or "?"
        by_bot[b] = by_bot.get(b, 0) + 1
        if r.get("feedbacks_text"):
            feedback_count += 1
        if r.get("persona_overridden"):
            persona_overridden_count += 1
        if r.get("crisis_trigger"):
            crisis_count += 1

    avg_response = (
        sum(r.get("response_seconds") or 0.0 for r in rows) / max(total, 1)
    )
    avg_answer_len = sum(len(r.get("assistant_answer") or "") for r in rows) / max(total, 1)

    meta = [
        ["항목", "값"],
        ["수집 시각 (UTC)", time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())],
        ["환경", "Production (Neon PostgreSQL)"],
        ["동일 질문 정책", "정규화(trim + lowercase) 후 가장 최근 답변만 유지 (DISTINCT ON)"],
        ["총 dedup 페어", total],
        ["피드백 달린 페어", feedback_count],
        ["페르소나 자동 override 페어", persona_overridden_count],
        ["위기 트리거 페어", crisis_count],
        ["평균 응답 시간(s)", round(avg_response, 1)],
        ["평균 답변 길이(자)", int(avg_answer_len)],
    ]
    meta.append(["", ""])
    meta.append(["챗봇별 페어 수", "(많은 순)"])
    for bot, n in sorted(by_bot.items(), key=lambda x: -x[1]):
        meta.append([bot, n])

    for r in meta:
        ws_meta.append(r)
    for col_idx in range(1, 3):
        ws_meta.cell(row=1, column=col_idx).fill = header_fill
        ws_meta.cell(row=1, column=col_idx).font = header_font
    ws_meta.column_dimensions["A"].width = 38
    ws_meta.column_dimensions["B"].width = 60
    for row in ws_meta.iter_rows(min_row=2, max_row=ws_meta.max_row, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"\n📊 저장 완료: {path}")
    print(f"   총 dedup 페어 {total:,} / 피드백 {feedback_count} / 평균 응답 {round(avg_response, 1)}s")

    # 요약 json도 별도
    OUT_SUMMARY.write_text(
        json.dumps(
            {
                "total_pairs": total,
                "feedback_count": feedback_count,
                "persona_overridden_count": persona_overridden_count,
                "crisis_count": crisis_count,
                "avg_response_seconds": round(avg_response, 1),
                "avg_answer_chars": int(avg_answer_len),
                "by_chatbot": dict(sorted(by_bot.items(), key=lambda x: -x[1])),
                "collected_at_utc": time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime()),
                "policy": "DISTINCT ON normalized question — keep most recent answer",
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    print(f"   summary: {OUT_SUMMARY}")


def main() -> int:
    rows = asyncio.run(fetch_rows())
    write_xlsx(rows, OUT_XLSX)
    return 0


if __name__ == "__main__":
    sys.exit(main())
