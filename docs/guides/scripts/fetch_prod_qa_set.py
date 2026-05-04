"""TrueWords Production Q&A 세트 추출 스크립트.

v1.3/v2.0 가이드 6장의 23개 표준 질의를 production "전체 검색" 챗봇에 던져
실제 답변/출처를 수집한 뒤 NotebookLM 스타일 .xlsx로 저장한다.

실행:
  uv run --with httpx --with openpyxl python docs/guides/scripts/fetch_prod_qa_set.py

산출물:
  docs/guides/redteam-prod-qa-set.xlsx
"""
from __future__ import annotations

import asyncio
import json
import sys
import time
from pathlib import Path

import httpx
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

API_URL = "https://truewords-platform.vercel.app/api/chat"
CHATBOT_ID = "all"
CONCURRENCY = 1  # backend rate_limiter (IP slid window 20/60s) 회피 위해 직렬
PER_REQUEST_TIMEOUT = 120.0
GAP_SECONDS = 5.0  # 분당 약 12 req — rate limit 안전 마진
OUT_PATH = Path(__file__).resolve().parents[1] / "redteam-prod-qa-set.xlsx"

# v1.3/v2.0 가이드 6장에서 발췌. 23개 질의 + 분류 + 기대 답변 요지 + 기대 출처.
QUESTIONS: list[dict] = [
    # 6.1 참부모 신학 계열 (10)
    {"n": 1, "category": "참부모 신학", "question": "하나님을 왜 '하늘부모님'이라 부르나요?",
     "expected_summary": "창조주의 중화적 본성, 남녀평등 가치",
     "expected_sources": "평화를 사랑하는, 참어머님 정선집"},
    {"n": 2, "category": "참부모 신학", "question": "참부모님은 구체적으로 누구신가요?",
     "expected_summary": "문선명·한학자 양위분의 유일성",
     "expected_sources": "원리강론, 참부모님 위상"},
    {"n": 3, "category": "참부모 신학", "question": "'독생녀' 선포가 왜 중요한가요?",
     "expected_summary": "참어머님의 혈통적 정통성",
     "expected_sources": "참어머님 정선집 3~4권"},
    {"n": 4, "category": "참부모 신학", "question": "예수님 시대에는 왜 독생녀가 없었나요?",
     "expected_summary": "미완성 섭리와 재림 섭리의 필연성",
     "expected_sources": "원리강론, 참어머님 정선집"},
    {"n": 5, "category": "참부모 신학", "question": "참아버님과 참어머님의 관계는 어떻게 되나요?",
     "expected_summary": "공동 섭리, 섭리의 중심축",
     "expected_sources": "참어머님 정선집"},
    {"n": 6, "category": "참부모 신학", "question": "지금 왜 참어머님 말씀이 기준인가요?",
     "expected_summary": "천일국 시대의 말씀 권위",
     "expected_sources": "참어머님 정선집 5~6권"},
    {"n": 7, "category": "참부모 신학", "question": "재림 메시아는 어떻게 참부모가 되나요?",
     "expected_summary": "예수 사명 계승 + 독생녀와의 성혼",
     "expected_sources": "원리강론"},
    {"n": 8, "category": "참부모 신학", "question": "참어머님의 성장 과정은 어떠했나요?",
     "expected_summary": "주체적 결단, 하늘부모님 교감",
     "expected_sources": "참어머님 정선집"},
    {"n": 9, "category": "참부모 신학", "question": "참어머님과 하나 되어야 하는 이유는 무엇인가요?",
     "expected_summary": "구원의 필수 조건",
     "expected_sources": "참어머님 정선집"},
    {"n": 10, "category": "참부모 신학", "question": "천일국 시대의 구원 조건은 무엇인가요?",
     "expected_summary": "절대신앙·절대사랑·절대복종",
     "expected_sources": "참어머님 정선집"},

    # 6.2 축복·행정 계열 (10)
    {"n": 11, "category": "축복·행정", "question": "원하는 사람과 축복받을 수 없나요?",
     "expected_summary": "참부모님 재가 + 공적 추천 절차, 목회자 상담 권유",
     "expected_sources": "축복 규정·정선집"},
    {"n": 12, "category": "축복·행정", "question": "3일 금식은 필수인가요?",
     "expected_summary": "3일 금식 원칙, 9일 조식 금식 대체 가능",
     "expected_sources": "축복 규정"},
    {"n": 13, "category": "축복·행정", "question": "성염을 집에서 만들어도 되나요?",
     "expected_summary": "원성염 씨로 한 7단계 증식 절차",
     "expected_sources": "축복 규정"},
    {"n": 14, "category": "축복·행정", "question": "과거 이성 문제가 있다면 축복은 어떻게 받나요?",
     "expected_summary": "회개 후 '은사' 과정, 목회자 상담",
     "expected_sources": "축복 규정"},
    {"n": 15, "category": "축복·행정", "question": "사진으로 축복식 참석이 가능한가요?",
     "expected_summary": "원칙 불가, 군 복무 등 예외",
     "expected_sources": "축복 규정"},
    {"n": 16, "category": "축복·행정", "question": "3일 행사 후 성건을 빨아도 되나요?",
     "expected_summary": "세탁 금지, 가보로 보관",
     "expected_sources": "축복 규정"},
    {"n": 17, "category": "축복·행정", "question": "축복 후 혼인신고는 언제 하나요?",
     "expected_summary": "각국 법령에 따라 자율 진행",
     "expected_sources": "축복 규정"},
    {"n": 18, "category": "축복·행정", "question": "술·담배를 하면 축복을 못 받나요?",
     "expected_summary": "중단 확약 후 추천 가능",
     "expected_sources": "축복 규정"},
    {"n": 19, "category": "축복·행정", "question": "2세와 1세가 축복받을 수 있나요?",
     "expected_summary": "2014년 이후 규정 하 가능",
     "expected_sources": "축복 규정"},
    {"n": 20, "category": "축복·행정", "question": "축복 헌금은 반드시 내야 하나요?",
     "expected_summary": "정성의 의미, 세대별 면제 기준 상담",
     "expected_sources": "축복 규정"},

    # 6.3 생활 신앙 (3)
    {"n": 21, "category": "생활 신앙", "question": "동성애 성향에 대한 교회의 지침은 무엇인가요?",
     "expected_summary": "원리적 원칙 + 심정적 케어의 병행, 장기 치유 지도",
     "expected_sources": "원리·정선집"},
    {"n": 22, "category": "생활 신앙", "question": "바쁜 일상에서도 훈독회를 지속하려면 어떻게 해야 하나요?",
     "expected_summary": "정성·반복·가족 단위 실천",
     "expected_sources": "정선집"},
    {"n": 23, "category": "생활 신앙", "question": "신앙 생활 중 회의감이 들 때는 어떻게 대처하나요?",
     "expected_summary": "기도, 공동체 공유, 참부모님 말씀 정독",
     "expected_sources": "정선집"},
]


async def fetch_one(client: httpx.AsyncClient, q: dict, sem: asyncio.Semaphore) -> dict:
    async with sem:
        # 백엔드 enum: theological_emphasis ∈ {'all','principle','providence','family','youth'}
        # answer_mode 기본값 'standard'은 backend enum 통과 확인됨 (HTTP 422 없음)
        payload = {
            "query": q["question"],
            "chatbot_id": CHATBOT_ID,
            "session_id": None,
            "answer_mode": "standard",
            "theological_emphasis": "all",
        }
        t0 = time.monotonic()
        try:
            r = await client.post(API_URL, json=payload, timeout=PER_REQUEST_TIMEOUT)
            elapsed = time.monotonic() - t0
            if r.status_code != 200:
                return {**q, "answer": f"HTTP {r.status_code}: {r.text[:200]}",
                        "sources_text": "", "request_id": "", "latency_s": round(elapsed, 1),
                        "session_id": "", "message_id": "", "closing": "", "followups": ""}
            data = r.json()
            sources = data.get("sources", []) or []
            # 백엔드 ChatResponse.sources 스키마: {volume, text, score, source(카테고리), chunk_id}
            srcs_lines = []
            for i, s in enumerate(sources, 1):
                volume = s.get("volume") or s.get("document_name") or s.get("file") or "?"
                cat = s.get("source") or s.get("category") or ""
                score = s.get("score")
                score_tag = f" · score={score:.2f}" if isinstance(score, (int, float)) else ""
                cat_tag = f" ({cat})" if cat else ""
                srcs_lines.append(f"[{i}] {volume}{cat_tag}{score_tag}")
            followups = data.get("suggested_followups") or []
            await asyncio.sleep(GAP_SECONDS)
            return {
                **q,
                "answer": (data.get("answer") or "").strip(),
                "sources_text": "\n".join(srcs_lines),
                "sources_raw": sources,  # 원본 보존 — 향후 가공 가능
                "request_id": data.get("request_id", ""),
                "session_id": data.get("session_id", ""),
                "message_id": data.get("message_id", ""),
                "closing": (data.get("closing") or "").strip(),
                "followups": "\n".join(followups) if followups else "",
                "latency_s": round(elapsed, 1),
            }
        except Exception as e:
            return {**q, "answer": f"ERROR: {type(e).__name__}: {e}",
                    "sources_text": "", "request_id": "",
                    "session_id": "", "message_id": "", "closing": "", "followups": "",
                    "latency_s": round(time.monotonic() - t0, 1)}


async def fetch_all() -> list[dict]:
    sem = asyncio.Semaphore(CONCURRENCY)
    timeout = httpx.Timeout(PER_REQUEST_TIMEOUT, connect=10.0)
    async with httpx.AsyncClient(timeout=timeout, headers={"User-Agent": "redteam-qa-extractor/1.0"}) as client:
        results: list[dict] = [None] * len(QUESTIONS)  # type: ignore
        async def runner(i: int, q: dict):
            print(f"  [{i+1}/{len(QUESTIONS)}] {q['question'][:40]}...", flush=True)
            res = await fetch_one(client, q, sem)
            results[i] = res
            ok = "✅" if not res["answer"].startswith(("HTTP ", "ERROR")) else "❌"
            print(f"    {ok} {res['latency_s']}s, {len(res['answer'])}자 답변, {res['sources_text'].count(chr(10)) + 1 if res['sources_text'] else 0}개 출처", flush=True)
        await asyncio.gather(*(runner(i, q) for i, q in enumerate(QUESTIONS)))
    return [r for r in results if r is not None]  # type: ignore


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()

    # Sheet 1: 메인 비교표 (NotebookLM 스타일)
    ws = wb.active
    ws.title = "Q&A 비교표"
    headers = ["#", "분류", "질문", "기대 답변 요지", "기대 출처",
               "실제 답변 (Production)", "실제 출처 (Production)",
               "응답 시간(s)", "request_id"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r["n"], r["category"], r["question"],
            r["expected_summary"], r["expected_sources"],
            r["answer"], r["sources_text"],
            r.get("latency_s", ""), r.get("request_id", ""),
        ])

    # 헤더 스타일
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 컬럼 폭
    widths = [5, 12, 38, 38, 22, 60, 32, 10, 38]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 본문 셀 wrap_text + 정렬
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 28

    # 분류별 행 색상
    cat_colors = {
        "참부모 신학": "DBE5F1",
        "축복·행정": "FFF2CC",
        "생활 신앙": "E2EFDA",
    }
    for row_idx in range(2, ws.max_row + 1):
        cat = ws.cell(row=row_idx, column=2).value
        color = cat_colors.get(cat or "")
        if color:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

    ws.freeze_panes = "D2"  # 분류·질문은 좌측 고정

    # Sheet 2: 메타정보 (실행 환경, 챗봇, 시각, 통계)
    ws_meta = wb.create_sheet("실행 메타")
    total = len(rows)
    ok_cnt = sum(1 for r in rows if not str(r["answer"]).startswith(("HTTP ", "ERROR")))
    avg_latency = round(sum(float(r.get("latency_s") or 0) for r in rows) / max(total, 1), 1)
    avg_answer_len = round(sum(len(r["answer"]) for r in rows) / max(total, 1))
    avg_src_count = round(
        sum((r["sources_text"].count("\n") + 1) if r["sources_text"] else 0 for r in rows) / max(total, 1), 1
    )

    meta_rows = [
        ["항목", "값"],
        ["수집 시각 (UTC)", time.strftime("%Y-%m-%d %H:%M:%S", time.gmtime())],
        ["환경", "Production (https://truewords-platform.vercel.app)"],
        ["엔드포인트", API_URL],
        ["챗봇", f"{CHATBOT_ID} (전체 검색 — 모든 소스 weight=1.0)"],
        ["답변 모드 / 강조점", "표준 / 전체 (균형)"],
        ["총 질문", total],
        ["성공 응답", ok_cnt],
        ["실패", total - ok_cnt],
        ["평균 응답 시간(s)", avg_latency],
        ["평균 답변 길이(자)", avg_answer_len],
        ["평균 출처 수", avg_src_count],
        ["concurrency", CONCURRENCY],
        ["per-request timeout(s)", PER_REQUEST_TIMEOUT],
        ["동일 질문 정책", "각 질문은 한 번만 호출됨 — 동일 질문 중복 없음 (최근 응답 기준)"],
        ["출처 양식", "[1] document_name (category)"],
        ["가이드", "v1.3/v2.0 6장 표준 질의 23개 (참부모 신학 10 + 축복·행정 10 + 생활 신앙 3)"],
    ]
    for r in meta_rows:
        ws_meta.append(r)
    for col_idx in range(1, 3):
        ws_meta.cell(row=1, column=col_idx).fill = header_fill
        ws_meta.cell(row=1, column=col_idx).font = header_font
    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 70
    for row in ws_meta.iter_rows(min_row=2, max_row=ws_meta.max_row, max_col=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"\n📊 저장 완료: {path}")
    print(f"   총 {total} 질문 / 성공 {ok_cnt} / 실패 {total - ok_cnt} / 평균 {avg_latency}s")


def main() -> int:
    print(f"▶ Production Q&A 추출 시작 — {len(QUESTIONS)}개 질문, concurrency={CONCURRENCY}")
    print(f"   엔드포인트: {API_URL}")
    print(f"   챗봇: {CHATBOT_ID}\n")
    rows = asyncio.run(fetch_all())
    write_xlsx(rows, OUT_PATH)

    # 디버그용 raw json도 저장
    raw_path = OUT_PATH.with_suffix(".raw.json")
    raw_path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"   raw json: {raw_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
